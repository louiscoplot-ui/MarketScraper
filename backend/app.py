import os
import io
import json
import logging
import time

import threading
from datetime import datetime
from collections import Counter
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

from database import init_db, get_db, add_suburb, remove_suburb, get_suburbs, get_listings
from database import upsert_listing, mark_withdrawn, create_scrape_log, update_scrape_log, get_scrape_logs
from database import get_existing_urls, trim_sold_listings, cleanup_agent_entries, restore_false_withdrawn
from database import backup_db, get_price_changes, take_market_snapshot, get_market_snapshots
from scraper import scrape_suburb, debug_page, compare_suburb, debug_detail, verify_disappeared_listings

app = Flask(__name__)
CORS(app)

# Track active scraping jobs
scrape_jobs = {}  # suburb_id -> {status, progress, started_at}
scrape_cancel = set()  # suburb_ids to cancel


@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok', 'app': 'market-scraper'})


# --- SUBURB AUTOCOMPLETE ---

@app.route('/api/suburbs/search', methods=['GET'])
def search_suburbs():
    """Autocomplete for WA suburb names."""
    from wa_suburbs import WA_SUBURBS
    q = request.args.get('q', '').strip().lower()
    if not q:
        return jsonify([])
    matches = [s for s in WA_SUBURBS if s.lower().startswith(q)]
    if not matches:
        matches = [s for s in WA_SUBURBS if q in s.lower()]
    return jsonify(matches[:15])


# --- SUBURB MANAGEMENT ---

@app.route('/api/suburbs', methods=['GET'])
def list_suburbs():
    return jsonify(get_suburbs())


@app.route('/api/suburbs', methods=['POST'])
def create_suburb():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400

    suburb = add_suburb(name)
    if suburb is None:
        return jsonify({'error': 'Suburb already exists'}), 409

    return jsonify(suburb), 201


@app.route('/api/suburbs/<int:suburb_id>', methods=['DELETE'])
def delete_suburb(suburb_id):
    remove_suburb(suburb_id)
    return jsonify({'ok': True})


# --- LISTINGS ---

@app.route('/api/listings/<int:listing_id>', methods=['DELETE'])
def delete_listing(listing_id):
    """Manual delete of a single listing row — useful for cleaning stale
    withdrawn rows that the auto re-list detector couldn't match (e.g. the
    withdrawn row has 'Address not disclosed' so no address to normalise)."""
    conn = get_db()
    row = conn.execute("SELECT id, address, status FROM listings WHERE id = ?",
                       (listing_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Listing not found'}), 404
    conn.execute("DELETE FROM listings WHERE id = ?", (listing_id,))
    conn.commit()
    conn.close()
    return jsonify({'deleted': listing_id, 'address': row['address'], 'status': row['status']})


@app.route('/api/listings', methods=['GET'])
def list_listings():
    suburb_id = request.args.get('suburb_id', type=int)
    suburb_ids_str = request.args.get('suburb_ids', '')
    status = request.args.get('status')
    statuses_str = request.args.get('statuses', '')
    suburb_ids = None
    if suburb_ids_str:
        try:
            suburb_ids = [int(x) for x in suburb_ids_str.split(',') if x.strip()]
        except ValueError:
            pass
    statuses = None
    if statuses_str:
        statuses = [s.strip() for s in statuses_str.split(',') if s.strip()]
    return jsonify(get_listings(suburb_id=suburb_id, suburb_ids=suburb_ids, status=status, statuses=statuses))


@app.route('/api/report', methods=['GET'])
def market_report():
    """Generate market report stats for selected suburbs."""
    import re as _re

    suburb_ids_str = request.args.get('suburb_ids', '')
    suburb_ids = None
    if suburb_ids_str:
        try:
            suburb_ids = [int(x) for x in suburb_ids_str.split(',') if x.strip()]
        except ValueError:
            pass

    listings = get_listings(suburb_ids=suburb_ids)
    if not listings:
        return jsonify({'error': 'No listings found'}), 404

    def parse_price(price_text):
        if not price_text:
            return None
        m = _re.search(r'\$([\d,]+)', price_text.replace(' ', ''))
        if m:
            try:
                return int(m.group(1).replace(',', ''))
            except ValueError:
                return None
        return None

    def calc_dom(l):
        # Only count DOM when REIWA actually published a listing date —
        # first_seen would just fabricate a number based on when we first
        # scraped the listing.
        date_str = l.get('listing_date') or ''
        if not date_str:
            return None
        ddmm = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
        try:
            if ddmm:
                start = datetime(int(ddmm.group(3)), int(ddmm.group(2)), int(ddmm.group(1)))
            else:
                start = datetime.fromisoformat(date_str.replace('Z', ''))
        except (ValueError, TypeError):
            return None
        return max(0, (datetime.utcnow() - start).days)

    # Overall stats
    active = [l for l in listings if l['status'] == 'active']
    under_offer = [l for l in listings if l['status'] == 'under_offer']
    sold = [l for l in listings if l['status'] == 'sold']
    withdrawn = [l for l in listings if l['status'] == 'withdrawn']

    # Price stats (active only)
    prices = [p for p in (parse_price(l.get('price_text')) for l in active) if p and p >= 100000]
    prices.sort()

    # DOM stats (active only)
    doms = [d for d in (calc_dom(l) for l in active) if d is not None]
    doms.sort()

    # Stale listings (60+ DOM)
    stale = [l for l in active if (calc_dom(l) or 0) >= 60]

    # Agent breakdown
    agent_stats = {}
    for l in listings:
        a = l.get('agent') or 'Unknown'
        if a not in agent_stats:
            agent_stats[a] = {'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0, 'total': 0}
        s = l.get('status', 'active')
        if s in agent_stats[a]:
            agent_stats[a][s] += 1
        agent_stats[a]['total'] += 1

    # Agency breakdown
    agency_stats = {}
    for l in listings:
        a = l.get('agency') or 'Unknown'
        if a not in agency_stats:
            agency_stats[a] = {'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0, 'total': 0}
        s = l.get('status', 'active')
        if s in agency_stats[a]:
            agency_stats[a][s] += 1
        agency_stats[a]['total'] += 1

    # Suburb breakdown
    suburb_stats = {}
    for l in listings:
        sn = l.get('suburb_name', 'Unknown')
        if sn not in suburb_stats:
            suburb_stats[sn] = {'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0, 'total': 0}
        s = l.get('status', 'active')
        if s in suburb_stats[sn]:
            suburb_stats[sn][s] += 1
        suburb_stats[sn]['total'] += 1

    # Property type breakdown
    type_stats = {}
    for l in active:
        t = l.get('listing_type') or 'Unknown'
        type_stats[t] = type_stats.get(t, 0) + 1

    # Market share: percentage of active listings per agency
    total_active = len(active)
    market_share = []
    if total_active > 0:
        agency_active = {}
        for l in active:
            a = l.get('agency') or 'Unknown'
            agency_active[a] = agency_active.get(a, 0) + 1
        market_share = sorted([
            {'agency': name, 'count': count, 'pct': round(count / total_active * 100, 1)}
            for name, count in agency_active.items()
        ], key=lambda x: x['count'], reverse=True)

    # Market share by suburb (for multi-suburb view)
    suburb_market_share = {}
    for l in active:
        sn = l.get('suburb_name', 'Unknown')
        a = l.get('agency') or 'Unknown'
        if sn not in suburb_market_share:
            suburb_market_share[sn] = {}
        suburb_market_share[sn][a] = suburb_market_share[sn].get(a, 0) + 1

    for sn in suburb_market_share:
        total_in_suburb = sum(suburb_market_share[sn].values())
        suburb_market_share[sn] = sorted([
            {'agency': name, 'count': count, 'pct': round(count / total_in_suburb * 100, 1)}
            for name, count in suburb_market_share[sn].items()
        ], key=lambda x: x['count'], reverse=True)

    # Price changes
    price_changes = get_price_changes(suburb_ids=suburb_ids, limit=30)
    price_drops = []
    for pc in price_changes:
        old_p = parse_price(pc.get('old_price'))
        new_p = parse_price(pc.get('new_price'))
        drop_amount = None
        drop_pct = None
        if old_p and new_p and new_p < old_p:
            drop_amount = old_p - new_p
            drop_pct = round((drop_amount / old_p) * 100, 1)
        price_drops.append({
            'address': pc.get('address'),
            'suburb': pc.get('suburb_name'),
            'old_price': pc.get('old_price'),
            'new_price': pc.get('new_price'),
            'drop_amount': drop_amount,
            'drop_pct': drop_pct,
            'changed_at': pc.get('changed_at'),
            'agent': pc.get('agent'),
            'agency': pc.get('agency'),
            'status': pc.get('status'),
            'reiwa_url': pc.get('reiwa_url'),
        })

    # Historical snapshots
    snapshots = get_market_snapshots(suburb_ids=suburb_ids, limit=90)

    report = {
        'generated_at': datetime.utcnow().isoformat(),
        'total_listings': len(listings),
        'summary': {
            'active': len(active),
            'under_offer': len(under_offer),
            'sold': len(sold),
            'withdrawn': len(withdrawn),
        },
        'price': {
            'count_with_price': len(prices),
            'min': min(prices) if prices else None,
            'max': max(prices) if prices else None,
            'median': prices[len(prices)//2] if prices else None,
            'avg': round(sum(prices) / len(prices)) if prices else None,
        },
        'dom': {
            'count': len(doms),
            'min': min(doms) if doms else None,
            'max': max(doms) if doms else None,
            'median': doms[len(doms)//2] if doms else None,
            'avg': round(sum(doms) / len(doms)) if doms else None,
            'stale_count': len(stale),
        },
        'market_share': market_share,
        'suburb_market_share': suburb_market_share,
        'price_drops': price_drops,
        'snapshots': snapshots,
        'stale_listings': [{
            'address': l.get('address'),
            'suburb': l.get('suburb_name'),
            'price': l.get('price_text'),
            'agent': l.get('agent'),
            'agency': l.get('agency'),
            'dom': calc_dom(l),
            'listing_date': l.get('listing_date'),
            'reiwa_url': l.get('reiwa_url'),
        } for l in sorted(stale, key=lambda x: calc_dom(x) or 0, reverse=True)],
        'agents': sorted(agent_stats.items(), key=lambda x: x[1]['total'], reverse=True),
        'agencies': sorted(agency_stats.items(), key=lambda x: x[1]['total'], reverse=True),
        'suburbs': sorted(suburb_stats.items(), key=lambda x: x[1]['total'], reverse=True),
        'property_types': sorted(type_stats.items(), key=lambda x: x[1], reverse=True),
        'withdrawn_listings': [{
            'address': l.get('address'),
            'suburb': l.get('suburb_name'),
            'price': l.get('price_text'),
            'agent': l.get('agent'),
            'agency': l.get('agency'),
            'listing_date': l.get('listing_date'),
            'reiwa_url': l.get('reiwa_url'),
        } for l in withdrawn],
    }

    return jsonify(report)


@app.route('/api/listings/export', methods=['GET'])
def export_listings():
    """Export filtered listings to a polished Excel workbook with native Tables.

    Features: sortable/filterable Excel Tables on every sheet, typed cells
    (Price as currency, Listed as date, Land/Internal as numeric m²), clickable
    REIWA hyperlinks, colored status text, DOM stale highlight, totals rows,
    freeze panes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import re as _re

    def _calc_dom(listing):
        # Only count DOM when REIWA published a listing date; don't fall back
        # to first_seen because that invents a DOM measured from when we
        # started scraping, not from the real listing day.
        date_str = listing.get('listing_date') or ''
        if not date_str:
            return None
        ddmm = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
        try:
            if ddmm:
                start = datetime(int(ddmm.group(3)), int(ddmm.group(2)), int(ddmm.group(1)))
            else:
                start = datetime.fromisoformat(date_str.replace('Z', ''))
        except (ValueError, TypeError):
            return None
        return max(0, (datetime.utcnow() - start).days)

    def _parse_price_numeric(text):
        """Extract a plausible AUD value from REIWA's free-form price string."""
        if not text:
            return None
        s = str(text).replace(',', '').replace(' ', '')
        m = _re.search(r'\$?(\d+(?:\.\d+)?)\s*([MmKk]?)', s)
        if not m:
            return None
        try:
            val = float(m.group(1))
        except ValueError:
            return None
        suffix = (m.group(2) or '').lower()
        if suffix == 'm':
            val *= 1_000_000
        elif suffix == 'k':
            val *= 1_000
        return int(val) if val >= 10_000 else None

    def _parse_size_numeric(text):
        if not text:
            return None
        m = _re.search(r'(\d[\d,]*)', str(text))
        if not m:
            return None
        try:
            return int(m.group(1).replace(',', ''))
        except ValueError:
            return None

    def _parse_listing_date(text):
        if not text:
            return None
        m = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', str(text).strip())
        if not m:
            return None
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
        except ValueError:
            return None

    STATUS_COLORS = {
        'active':      '15803D',  # green
        'under_offer': 'B45309',  # amber
        'sold':        '4B5563',  # slate
        'withdrawn':   'B91C1C',  # red
    }

    # Palette — dark forest green header, warm gold totals, mint banding
    HEADER_FILL = PatternFill('solid', fgColor='1E3A2B')
    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BAND_FILL = PatternFill('solid', fgColor='F4F7F5')
    TOTAL_FILL = PatternFill('solid', fgColor='F5E6A8')
    TOTAL_FONT = Font(name='Calibri', bold=True, color='1E3A2B', size=11)
    BODY_FONT = Font(name='Calibri', size=10)
    BODY_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', indent=1)
    BODY_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    BODY_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', indent=1)
    SUBTLE_BORDER = Border(bottom=Side(style='thin', color='D4E0D8'))

    # Light1 means "no extra styling" — our custom fills come through cleanly;
    # we keep the Table purely for the per-column filter/sort buttons in Excel.
    TABLE_STYLE = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                 showColumnStripes=False, showFirstColumn=False,
                                 showLastColumn=False)

    def _style_header_row(sheet, n_cols, height=30):
        sheet.row_dimensions[1].height = height
        for c in range(1, n_cols + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = SUBTLE_BORDER

    # Parse same filters as list_listings
    suburb_ids_str = request.args.get('suburb_ids', '')
    statuses_str = request.args.get('statuses', '')
    agent = request.args.get('agent', '').strip()
    agency = request.args.get('agency', '').strip()

    suburb_ids = None
    if suburb_ids_str:
        try:
            suburb_ids = [int(x) for x in suburb_ids_str.split(',') if x.strip()]
        except ValueError:
            pass
    statuses = None
    if statuses_str:
        statuses = [s.strip() for s in statuses_str.split(',') if s.strip()]

    listings = get_listings(suburb_ids=suburb_ids, statuses=statuses)
    if agent:
        listings = [l for l in listings if l.get('agent') == agent]
    if agency:
        listings = [l for l in listings if l.get('agency') == agency]

    # Default order: newest listing_date first
    listings = sorted(
        listings,
        key=lambda l: _parse_listing_date(l.get('listing_date')) or datetime.min.date(),
        reverse=True,
    )

    wb = Workbook()

    # === Sheet 1: Listings ===
    ws = wb.active
    ws.title = "Listings"
    columns = ['Address', 'Suburb', 'Price', 'Price (AUD)', 'Bed', 'Bath', 'Car',
               'Land (m²)', 'Internal (m²)', 'Agency', 'Agent', 'Listed', 'DOM',
               'Withdrawn', 'Status', 'Type', 'Link']
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    stale_font = Font(name='Calibri', bold=True, color='CC0000', size=10)
    # Per-column alignment (1-indexed matching `columns`)
    col_align = {
        1: BODY_ALIGN_LEFT, 2: BODY_ALIGN_LEFT, 3: BODY_ALIGN_LEFT,
        4: BODY_ALIGN_RIGHT, 5: BODY_ALIGN_CENTER, 6: BODY_ALIGN_CENTER,
        7: BODY_ALIGN_CENTER, 8: BODY_ALIGN_RIGHT, 9: BODY_ALIGN_RIGHT,
        10: BODY_ALIGN_LEFT, 11: BODY_ALIGN_LEFT, 12: BODY_ALIGN_CENTER,
        13: BODY_ALIGN_CENTER, 14: BODY_ALIGN_CENTER, 15: BODY_ALIGN_CENTER,
        16: BODY_ALIGN_CENTER, 17: BODY_ALIGN_CENTER,
    }

    def _iso_to_date(iso):
        if not iso:
            return None
        try:
            return datetime.fromisoformat(iso.replace('Z', '').split('.')[0]).date()
        except (ValueError, AttributeError):
            return None

    for row_idx, l in enumerate(listings, 2):
        status_raw = (l.get('status') or '').lower()
        price_text = l.get('price_text') or ''
        price_num = _parse_price_numeric(price_text)
        land_num = _parse_size_numeric(l.get('land_size'))
        internal_num = _parse_size_numeric(l.get('internal_size'))
        listed_date = _parse_listing_date(l.get('listing_date'))
        url = l.get('reiwa_url') or ''
        dom = _calc_dom(l)
        row_fill = BAND_FILL if row_idx % 2 == 0 else None

        ws.cell(row=row_idx, column=1, value=l.get('address', ''))
        ws.cell(row=row_idx, column=2, value=l.get('suburb_name', ''))
        ws.cell(row=row_idx, column=3, value=price_text)
        pcell = ws.cell(row=row_idx, column=4, value=price_num)
        if price_num is not None:
            pcell.number_format = '"$"#,##0'
        ws.cell(row=row_idx, column=5, value=l.get('bedrooms'))
        ws.cell(row=row_idx, column=6, value=l.get('bathrooms'))
        ws.cell(row=row_idx, column=7, value=l.get('parking'))
        lcell = ws.cell(row=row_idx, column=8, value=land_num)
        if land_num is not None:
            lcell.number_format = '#,##0" m²"'
        icell = ws.cell(row=row_idx, column=9, value=internal_num)
        if internal_num is not None:
            icell.number_format = '#,##0" m²"'
        ws.cell(row=row_idx, column=10, value=l.get('agency', ''))
        ws.cell(row=row_idx, column=11, value=l.get('agent', ''))
        dcell = ws.cell(row=row_idx, column=12, value=listed_date)
        if listed_date:
            dcell.number_format = 'DD/MM/YYYY'
        dom_cell = ws.cell(row=row_idx, column=13, value=dom)
        withdrawn_date = _iso_to_date(l.get('withdrawn_date'))
        wcell = ws.cell(row=row_idx, column=14, value=withdrawn_date)
        if withdrawn_date:
            wcell.number_format = 'DD/MM/YYYY'
        scell = ws.cell(row=row_idx, column=15,
                        value=status_raw.replace('_', ' ').title() if status_raw else '')
        ws.cell(row=row_idx, column=16, value=l.get('listing_type', ''))
        link_cell = ws.cell(row=row_idx, column=17, value="View on REIWA" if url else '')
        if url:
            link_cell.hyperlink = url

        # Row-wide styling (font, alignment, banding, bottom border)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = BODY_FONT
            cell.alignment = col_align[c]
            cell.border = SUBTLE_BORDER
            if row_fill:
                cell.fill = row_fill

        # Per-cell overrides on top of the row styling
        if dom is not None and dom >= 60:
            dom_cell.font = stale_font
        color = STATUS_COLORS.get(status_raw)
        if color:
            scell.font = Font(name='Calibri', bold=True, color=color, size=10)
        if url:
            link_cell.font = Font(name='Calibri', color='1E3A8A', underline='single', size=10)

        ws.row_dimensions[row_idx].height = 22

    # Fixed widths tuned for the column contents
    #                  Addr Sub  Pri  P(AUD) B  B  C  Land Intl Agcy Agt  Lstd DOM Wdrn Sts Typ Link
    for col_idx, width in enumerate(
        [34, 18, 20, 14, 6, 6, 6, 13, 15, 28, 22, 12, 8, 12, 14, 14, 18], 1
    ):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    _style_header_row(ws, len(columns), height=34)
    ws.freeze_panes = "A2"

    last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
    last_row = max(2, len(listings) + 1)
    listings_table = Table(displayName="ListingsTable",
                           ref=f"A1:{last_col_letter}{last_row}")
    listings_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(listings_table)

    # === Summary sheet helper ===
    def _write_summary(sheet_name, table_name, headers, rows_data, col_widths):
        s = wb.create_sheet(sheet_name)
        for col_idx, h in enumerate(headers, 1):
            s.cell(row=1, column=col_idx, value=h)
        for row_idx, row_vals in enumerate(rows_data, 2):
            row_fill = BAND_FILL if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row_vals, 1):
                cell = s.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN_LEFT if col_idx == 1 else BODY_ALIGN_CENTER
                cell.border = SUBTLE_BORDER
                if row_fill:
                    cell.fill = row_fill
            s.row_dimensions[row_idx].height = 20
        # Totals row below the table (not inside the Table range)
        if rows_data:
            totals_row = len(rows_data) + 2
            lbl = s.cell(row=totals_row, column=1, value='Total')
            lbl.font = TOTAL_FONT
            lbl.fill = TOTAL_FILL
            lbl.alignment = BODY_ALIGN_LEFT
            for col_idx in range(2, len(headers) + 1):
                letter = s.cell(row=1, column=col_idx).column_letter
                cell = s.cell(row=totals_row, column=col_idx,
                              value=f"=SUM({letter}2:{letter}{totals_row - 1})")
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.alignment = BODY_ALIGN_CENTER
            s.row_dimensions[totals_row].height = 22
        for col_idx, w in enumerate(col_widths, 1):
            s.column_dimensions[s.cell(row=1, column=col_idx).column_letter].width = w
        _style_header_row(s, len(headers), height=30)
        s.freeze_panes = "A2"
        last_data_row = max(2, len(rows_data) + 1)
        last_letter = s.cell(row=1, column=len(headers)).column_letter
        tbl = Table(displayName=table_name, ref=f"A1:{last_letter}{last_data_row}")
        tbl.tableStyleInfo = TABLE_STYLE
        s.add_table(tbl)

    def _group_stats(listings_, key):
        buckets = {}
        for l in listings_:
            k = l.get(key) or 'Unknown'
            b = buckets.setdefault(k, {'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0, 'total': 0})
            b['total'] += 1
            st = (l.get('status') or 'active').lower()
            if st in b:
                b[st] += 1
        rows = [
            [k, v['total'], v['active'], v['under_offer'], v['sold'], v['withdrawn']]
            for k, v in buckets.items()
        ]
        rows.sort(key=lambda r: -r[1])
        return rows

    # === Sheet 2: Agents ===
    _write_summary('Agents', 'AgentsTable',
                   ['Agent', 'Total', 'Active', 'Under Offer', 'Sold', 'Withdrawn'],
                   _group_stats(listings, 'agent'),
                   [26, 10, 10, 14, 10, 12])

    # === Sheet 3: Agencies ===
    _write_summary('Agencies', 'AgenciesTable',
                   ['Agency', 'Total', 'Active', 'Under Offer', 'Sold', 'Withdrawn'],
                   _group_stats(listings, 'agency'),
                   [34, 10, 10, 14, 10, 12])

    # === Sheet 4: Suburb Summary (only if multiple suburbs) ===
    suburb_names = set(l.get('suburb_name', '') for l in listings)
    if len(suburb_names) > 1:
        suburb_data = {}
        for l in listings:
            sn = l.get('suburb_name', 'Unknown')
            d = suburb_data.setdefault(sn, {
                'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0,
                'agents': set(), 'agencies': set(),
            })
            st = (l.get('status') or 'active').lower()
            if st in d:
                d[st] += 1
            if l.get('agent'):
                d['agents'].add(l['agent'])
            if l.get('agency'):
                d['agencies'].add(l['agency'])
        suburb_rows = []
        for sname, data in sorted(suburb_data.items()):
            total = data['active'] + data['under_offer'] + data['sold'] + data['withdrawn']
            suburb_rows.append([
                sname, total, data['active'], data['under_offer'], data['sold'],
                data['withdrawn'], len(data['agents']), len(data['agencies'])
            ])
        _write_summary('Suburb Summary', 'SuburbsTable',
                       ['Suburb', 'Total', 'Active', 'Under Offer', 'Sold',
                        'Withdrawn', 'Agents', 'Agencies'],
                       suburb_rows,
                       [22, 10, 10, 14, 10, 12, 10, 12])

    # Save to buffer and send
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = datetime.now().strftime('%Y-%m-%d')
    suburb_label = ''
    if suburb_ids:
        conn = get_db()
        names = []
        for sid in suburb_ids:
            row = conn.execute("SELECT name FROM suburbs WHERE id = ?", (sid,)).fetchone()
            if row:
                names.append(row['name'])
        conn.close()
        if len(names) <= 3:
            suburb_label = '_' + '_'.join(names)
        else:
            suburb_label = f'_{len(names)}_suburbs'

    filename = f'MarketScraper{suburb_label}_{date_str}.xlsx'

    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/listings/summary', methods=['GET'])
def listings_summary():
    """Get a summary of listings counts by suburb and status."""
    conn = get_db()
    rows = conn.execute("""
        SELECT
            s.id as suburb_id,
            s.name as suburb_name,
            l.status,
            COUNT(*) as count
        FROM listings l
        JOIN suburbs s ON l.suburb_id = s.id
        GROUP BY s.id, l.status
        ORDER BY s.name
    """).fetchall()
    conn.close()

    summary = {}
    for row in rows:
        sid = row['suburb_id']
        if sid not in summary:
            summary[sid] = {
                'suburb_id': row['suburb_id'],
                'suburb_name': row['suburb_name'],
                'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0
            }
        summary[sid][row['status']] = row['count']

    return jsonify(list(summary.values()))


# --- SCRAPING ---

@app.route('/api/scrape/<int:suburb_id>', methods=['POST'])
def start_scrape(suburb_id):
    """Start scraping a suburb. Runs in background thread."""
    if suburb_id in scrape_jobs and scrape_jobs[suburb_id].get('status') == 'running':
        return jsonify({'error': 'Scrape already in progress for this suburb'}), 409

    conn = get_db()
    suburb = conn.execute("SELECT * FROM suburbs WHERE id = ?", (suburb_id,)).fetchone()
    conn.close()

    if not suburb:
        return jsonify({'error': 'Suburb not found'}), 404

    scrape_jobs[suburb_id] = {
        'status': 'running',
        'progress': 'Starting...',
        'started_at': datetime.utcnow().isoformat(),
    }

    thread = threading.Thread(
        target=_run_scrape,
        args=(suburb_id, suburb['slug'], suburb['name']),
        daemon=True
    )
    thread.start()

    return jsonify({'status': 'started', 'suburb': suburb['name']})


@app.route('/api/scrape/all', methods=['POST'])
def start_scrape_all():
    """Start scraping all active suburbs sequentially."""
    suburbs = get_suburbs()
    active_suburbs = [s for s in suburbs if s['active']]

    if not active_suburbs:
        return jsonify({'error': 'No active suburbs'}), 400

    # Check if any scrape is already running
    for s in active_suburbs:
        if s['id'] in scrape_jobs and scrape_jobs[s['id']].get('status') == 'running':
            return jsonify({'error': f'Scrape already running for {s["name"]}'}), 409

    thread = threading.Thread(
        target=_run_scrape_all,
        args=(active_suburbs,),
        daemon=True
    )
    thread.start()

    return jsonify({'status': 'started', 'suburbs': [s['name'] for s in active_suburbs]})


@app.route('/api/scrape/cancel', methods=['POST'])
def cancel_scrape():
    """Cancel all running scrapes."""
    running_ids = [sid for sid, job in scrape_jobs.items() if job.get('status') == 'running']
    for sid in running_ids:
        scrape_cancel.add(sid)
        scrape_jobs[sid]['progress'] = 'Cancelling...'
    logger.info(f"Cancel requested for {len(running_ids)} suburb(s): {running_ids}")
    return jsonify({'cancelled': running_ids})


@app.route('/api/scrape/status', methods=['GET'])
def scrape_status():
    """Get status of all scrape jobs."""
    return jsonify(scrape_jobs)


@app.route('/api/scrape/status/<int:suburb_id>', methods=['GET'])
def scrape_status_single(suburb_id):
    job = scrape_jobs.get(suburb_id, {'status': 'idle'})
    return jsonify(job)


@app.route('/api/scrape/debug/<int:suburb_id>', methods=['GET'])
def debug_scrape(suburb_id):
    """Debug: see what the scraper sees on the REIWA page for a suburb."""
    conn = get_db()
    suburb = conn.execute("SELECT * FROM suburbs WHERE id = ?", (suburb_id,)).fetchone()
    conn.close()
    if not suburb:
        return jsonify({'error': 'Suburb not found'}), 404
    result = debug_page(suburb['slug'])
    return jsonify(result)


@app.route('/api/admin/reset-listing-dates', methods=['POST'])
def reset_listing_dates():
    """Clear listing_date on all rows so the next scrape repopulates them.

    Needed once after the regex broadening introduced a false-positive that
    set many listings to today's date — NULLIF upsert would otherwise
    preserve those wrong values forever.
    """
    conn = get_db()
    cur = conn.execute("UPDATE listings SET listing_date = NULL WHERE listing_date IS NOT NULL")
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return jsonify({'cleared': affected})


@app.route('/api/scrape/debug-detail', methods=['GET'])
def debug_scrape_detail():
    """Debug a single listing URL: returns extracted fields, text snippets,
    and regex-match results so we can see why land/internal sizes are empty.

    Usage: /api/scrape/debug-detail?url=https://reiwa.com.au/28-forrest-street-cottesloe-5011589
    """
    url = request.args.get('url', '').strip()
    if not url:
        return jsonify({'error': 'Missing ?url=...'}), 400
    return jsonify(debug_detail(url))


@app.route('/api/scrape/compare/<int:suburb_id>', methods=['GET'])
def compare_scrape(suburb_id):
    """Compare REIWA's live listings vs our DB for a suburb."""
    conn = get_db()
    suburb = conn.execute("SELECT * FROM suburbs WHERE id = ?", (suburb_id,)).fetchone()
    if not suburb:
        conn.close()
        return jsonify({'error': 'Suburb not found'}), 404
    # Get all active/under_offer URLs from DB
    rows = conn.execute(
        "SELECT reiwa_url FROM listings WHERE suburb_id = ? AND status IN ('active', 'under_offer')",
        (suburb_id,)
    ).fetchall()
    conn.close()
    db_urls = {r['reiwa_url'] for r in rows if r['reiwa_url']}
    result = compare_suburb(suburb['slug'], db_urls)
    result['suburb'] = suburb['name']
    return jsonify(result)


@app.route('/api/scrape/audit', methods=['GET'])
def audit_suburbs():
    """Multi-suburb audit — data completeness + optional REIWA comparison.

    Query params:
      suburb_ids=1,2,3   REQUIRED — comma-separated suburb IDs
      compare=true       OPTIONAL — if set, also fetch REIWA live and diff URLs
                         (slow: adds ~10-15s per suburb)

    Returns per suburb:
      - db_count               total active/under_offer rows in DB
      - reiwa_total, matched, missing_from_db, sold_excluded  (compare only)
      - completeness           counts of rows missing each important field
      - incomplete_examples    up to 10 sample rows for diagnosis
    """
    ids_str = request.args.get('suburb_ids', '').strip()
    if not ids_str:
        return jsonify({'error': 'suburb_ids required (comma-separated)'}), 400
    try:
        suburb_ids = [int(x) for x in ids_str.split(',') if x.strip()]
    except ValueError:
        return jsonify({'error': 'invalid suburb_ids'}), 400
    do_compare = request.args.get('compare', '').lower() in ('1', 'true', 'yes')

    conn = get_db()
    results = []

    STRATA_TYPES = {'unit', 'apartment', 'townhouse', 'villa', 'studio', 'duplex'}

    for sid in suburb_ids:
        suburb = conn.execute("SELECT * FROM suburbs WHERE id = ?", (sid,)).fetchone()
        if not suburb:
            results.append({'suburb_id': sid, 'error': 'Suburb not found'})
            continue

        rows = conn.execute(
            """
            SELECT address, reiwa_url, listing_type, land_size, internal_size,
                   price_text, agent, agency, bedrooms, bathrooms, listing_date
            FROM listings
            WHERE suburb_id = ? AND status IN ('active', 'under_offer')
            """,
            (sid,)
        ).fetchall()

        # Completeness tally — each listing counted AT MOST ONCE per bucket
        missing_land = []
        missing_internal = []
        missing_type = []
        missing_price = []
        missing_agent = []
        missing_agency = []
        missing_date = []
        missing_beds = []
        for r in rows:
            t = (r['listing_type'] or '').strip().lower()
            addr = r['address'] or '(no address)'
            url = r['reiwa_url']
            land = (r['land_size'] or '').strip()
            internal = (r['internal_size'] or '').strip()

            # Land only required for houses (and for unknown types with no sizes
            # at all — those are suspicious, count them once here).
            if (t == 'house' and not land) or (not t and not land and not internal):
                missing_land.append({'address': addr, 'url': url, 'type': t or '(unknown)'})
            # Internal only required for strata-style dwellings.
            if t in STRATA_TYPES and not internal:
                missing_internal.append({'address': addr, 'url': url, 'type': t})
            if not t:
                missing_type.append({'address': addr, 'url': url})
            if not (r['price_text'] or '').strip():
                missing_price.append({'address': addr, 'url': url})
            if not (r['agent'] or '').strip():
                missing_agent.append({'address': addr, 'url': url})
            if not (r['agency'] or '').strip():
                missing_agency.append({'address': addr, 'url': url})
            if not (r['listing_date'] or '').strip():
                missing_date.append({'address': addr, 'url': url})
            if r['bedrooms'] is None:
                missing_beds.append({'address': addr, 'url': url})

        entry = {
            'suburb_id': sid,
            'suburb': suburb['name'],
            'db_count': len(rows),
            'completeness': {
                'missing_land_size': len(missing_land),
                'missing_internal_size': len(missing_internal),
                'missing_listing_type': len(missing_type),
                'missing_price': len(missing_price),
                'missing_agent': len(missing_agent),
                'missing_agency': len(missing_agency),
                'missing_listing_date': len(missing_date),
                'missing_bedrooms': len(missing_beds),
            },
            'examples': {
                'missing_land': missing_land[:10],
                'missing_internal': missing_internal[:10],
                'missing_type': missing_type[:10],
                'missing_price': missing_price[:10],
            },
        }

        if do_compare:
            db_urls = {r['reiwa_url'] for r in rows if r['reiwa_url']}
            try:
                cmp_result = compare_suburb(suburb['slug'], db_urls)
                entry['reiwa_total'] = cmp_result.get('reiwa_total')
                entry['matched'] = cmp_result.get('matched')
                entry['missing_from_db'] = cmp_result.get('missing_from_db', [])
                entry['sold_excluded'] = cmp_result.get('sold_excluded', [])
                entry['extra_in_db'] = cmp_result.get('extra_in_db', [])
                entry['pages_scraped'] = cmp_result.get('pages_scraped')
            except Exception as e:
                entry['compare_error'] = str(e)

        results.append(entry)

    conn.close()
    return jsonify({
        'suburbs': results,
        'compare_mode': do_compare,
    })


@app.route('/api/scrape/selected', methods=['POST'])
def scrape_selected():
    """Scrape only selected suburb IDs."""
    data = request.json
    suburb_ids = data.get('suburb_ids', [])
    if not suburb_ids:
        return jsonify({'error': 'No suburbs selected'}), 400

    conn = get_db()
    suburbs_to_scrape = []
    for sid in suburb_ids:
        s = conn.execute("SELECT * FROM suburbs WHERE id = ?", (sid,)).fetchone()
        if s:
            suburbs_to_scrape.append(dict(s))
    conn.close()

    if not suburbs_to_scrape:
        return jsonify({'error': 'No valid suburbs found'}), 400

    for s in suburbs_to_scrape:
        if s['id'] in scrape_jobs and scrape_jobs[s['id']].get('status') == 'running':
            return jsonify({'error': f'Scrape already running for {s["name"]}'}), 409

    thread = threading.Thread(
        target=_run_scrape_all,
        args=([{'id': s['id'], 'slug': s['slug'], 'name': s['name']} for s in suburbs_to_scrape],),
        daemon=True
    )
    thread.start()

    return jsonify({'status': 'started', 'suburbs': [s['name'] for s in suburbs_to_scrape]})


@app.route('/api/scrape/logs', methods=['GET'])
def list_scrape_logs():
    suburb_id = request.args.get('suburb_id', type=int)
    return jsonify(get_scrape_logs(suburb_id=suburb_id))




def _run_scrape_all(suburbs):
    """Run scrape for suburbs in parallel (up to 8 at a time)."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    max_workers = min(8, len(suburbs))

    # Mark all suburbs as queued first
    for s in suburbs:
        scrape_cancel.discard(s['id'])

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_run_scrape, s['id'], s['slug'], s['name']): s
            for s in suburbs
        }
        for future in as_completed(futures):
            s = futures[future]
            try:
                future.result()
            except Exception as e:
                logger.error(f"Scrape thread error for {s['name']}: {e}")


def _run_scrape(suburb_id, slug, name):
    """Execute the scraping process for a single suburb."""
    log_id = create_scrape_log(suburb_id)
    scrape_jobs[suburb_id] = {
        'status': 'running',
        'progress': f'Starting scrape for {name}...',
        'started_at': datetime.utcnow().isoformat(),
    }

    def progress_cb(msg):
        scrape_jobs[suburb_id]['progress'] = msg
        logger.info(f"[{name}] {msg}")

    try:
        # Get known URLs to skip detail pages for existing listings
        known_urls = get_existing_urls(suburb_id)
        logger.info(f"[{name}] {len(known_urls)} known URLs in DB, will skip their detail pages")

        def cancel_check():
            return suburb_id in scrape_cancel

        result = scrape_suburb(slug, suburb_id, progress_callback=progress_cb, known_urls=known_urls, cancel_check=cancel_check)

        # Check if cancelled
        if suburb_id in scrape_cancel:
            scrape_cancel.discard(suburb_id)
            scrape_jobs[suburb_id] = {
                'status': 'cancelled',
                'progress': 'Scrape cancelled by user',
                'completed_at': datetime.utcnow().isoformat(),
            }
            update_scrape_log(log_id, completed_at=datetime.utcnow().isoformat(), errors='Cancelled by user')
            logger.info(f"Scrape cancelled for {name}")
            return

        # Process for-sale listings — keyed by reiwa_url (no address dedup)
        # Same property by 2 agencies = 2 different REIWA URLs = 2 rows
        new_count = 0
        updated_count = 0
        forsale_urls = []
        sold_urls = []

        progress_cb('Saving for-sale listings to database...')
        for listing in result['forsale_listings']:
            url = listing.get('reiwa_url', '').strip()
            if not url:
                continue
            forsale_urls.append(url)
            action = upsert_listing(suburb_id, url, listing)
            if action == 'new':
                new_count += 1
            else:
                updated_count += 1

        # Process sold listings
        progress_cb('Saving sold listings to database...')
        saved_sold = 0
        for listing in result['sold_listings']:
            url = listing.get('reiwa_url', '').strip()
            if not url:
                continue
            sold_urls.append(url)
            upsert_listing(suburb_id, url, listing)
            saved_sold += 1

        # Pre-withdrawn verification — rescue listings that disappeared from
        # the for-sale grid AND fell past page 10 of sold (so they weren't in
        # sold_urls either). Visit each candidate's detail page; REIWA usually
        # still serves the page with a clear SOLD or UNDER OFFER badge.
        progress_cb('Verifying disappeared listings...')
        conn = get_db()
        db_active_rows = conn.execute(
            "SELECT reiwa_url FROM listings WHERE suburb_id = ? "
            "AND status IN ('active', 'under_offer') AND reiwa_url IS NOT NULL",
            (suburb_id,)
        ).fetchall()
        conn.close()
        db_active_urls = {r['reiwa_url'].rstrip('/') for r in db_active_rows}
        seen_set = {u.rstrip('/') for u in (forsale_urls + sold_urls)}
        candidates = list(db_active_urls - seen_set)

        rescued_sold = 0
        rescued_active = 0
        if candidates:
            logger.info(f"[{name}] Verifying {len(candidates)} disappeared URL(s) via detail pages")
            verify = verify_disappeared_listings(candidates)
            for url, status in verify.items():
                if status == 'sold':
                    upsert_listing(suburb_id, url, {'status': 'sold'})
                    sold_urls.append(url)
                    rescued_sold += 1
                elif status in ('active', 'under_offer'):
                    upsert_listing(suburb_id, url, {'status': status})
                    forsale_urls.append(url)
                    rescued_active += 1
                # 'gone' → leave it, the next mark_withdrawn call handles it
            if rescued_sold:
                logger.info(f"[{name}] Rescued {rescued_sold} from withdrawn → marked SOLD")
            if rescued_active:
                logger.info(f"[{name}] Rescued {rescued_active} from withdrawn → still active")

        # Mark withdrawn — only what's left after individual verification can
        # confidently be considered withdrawn, so we don't gate on coverage %.
        progress_cb('Checking for withdrawn listings...')
        reiwa_total = result['stats'].get('reiwa_total', 0)
        our_count = len(forsale_urls)
        if reiwa_total > 0:
            confident = our_count >= reiwa_total or (our_count >= reiwa_total * 0.95 and reiwa_total - our_count <= 3)
        else:
            confident = False
        # Verification tightens us up: any candidate still missing after a
        # successful detail-page round trip is genuinely withdrawn, regardless
        # of overall scrape coverage. Only skip the mark step if we had no
        # candidates to verify AND the scrape was incomplete.
        if candidates:
            confident = True
        if confident:
            logger.info(f"[{name}] Confident scrape (verified): {our_count} found vs {reiwa_total} REIWA total — checking withdrawals")
        else:
            logger.info(f"[{name}] Incomplete scrape: {our_count} found vs {reiwa_total} REIWA total — skipping withdrawals")
        withdrawn_count = mark_withdrawn(suburb_id, forsale_urls, sold_urls, confident=confident)

        # Keep only 40 most recent sold listings
        trimmed = trim_sold_listings(suburb_id, keep=40)
        if trimmed:
            logger.info(f"[{name}] Trimmed {trimmed} old sold listings (keeping 40 most recent)")

        # Use actual saved counts (not raw card counts)
        actual_forsale = len(forsale_urls)
        update_scrape_log(
            log_id,
            completed_at=datetime.utcnow().isoformat(),
            forsale_count=actual_forsale,
            sold_count=saved_sold,
            new_count=new_count,
            updated_count=updated_count,
            withdrawn_count=withdrawn_count,
            errors=json.dumps(result['errors']) if result['errors'] else None
        )

        # Take market snapshot for historical tracking
        try:
            import re as _re
            snap_conn = get_db()
            snap_rows = snap_conn.execute(
                "SELECT status, price_text, listing_date, first_seen FROM listings WHERE suburb_id = ?",
                (suburb_id,)
            ).fetchall()
            snap_conn.close()

            snap_active = [r for r in snap_rows if r['status'] == 'active']
            snap_uo = [r for r in snap_rows if r['status'] == 'under_offer']
            snap_sold = [r for r in snap_rows if r['status'] == 'sold']
            snap_wd = [r for r in snap_rows if r['status'] == 'withdrawn']

            # Median price
            snap_prices = []
            for r in snap_active:
                pt = r['price_text']
                if pt:
                    m = _re.search(r'\$([\d,]+)', pt.replace(' ', ''))
                    if m:
                        try:
                            p = int(m.group(1).replace(',', ''))
                            if p >= 100000:
                                snap_prices.append(p)
                        except ValueError:
                            pass
            snap_prices.sort()
            median_p = snap_prices[len(snap_prices)//2] if snap_prices else None

            # Avg DOM
            snap_doms = []
            for r in snap_active:
                ds = r['listing_date'] or ''
                dm = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', ds)
                try:
                    if dm:
                        start = datetime(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
                    else:
                        start = datetime.fromisoformat(ds.replace('Z', ''))
                    snap_doms.append(max(0, (datetime.utcnow() - start).days))
                except (ValueError, TypeError):
                    pass
            avg_d = round(sum(snap_doms) / len(snap_doms)) if snap_doms else None

            take_market_snapshot(suburb_id, {
                'active': len(snap_active),
                'under_offer': len(snap_uo),
                'sold': len(snap_sold),
                'withdrawn': len(snap_wd),
                'new': new_count,
                'median_price': median_p,
                'avg_dom': avg_d,
            })
            logger.info(f"[{name}] Market snapshot saved")
        except Exception as snap_err:
            logger.warning(f"[{name}] Failed to save snapshot: {snap_err}")

        scrape_jobs[suburb_id] = {
            'status': 'completed',
            'progress': f'Done! {actual_forsale} active, {saved_sold} sold, {new_count} new, {withdrawn_count} withdrawn',
            'completed_at': datetime.utcnow().isoformat(),
            'stats': result['stats'],
        }
        logger.info(f"Scrape completed for {name}: {result['stats']}")

    except Exception as e:
        logger.error(f"Scrape failed for {name}: {e}")
        update_scrape_log(log_id, completed_at=datetime.utcnow().isoformat(), errors=str(e))
        scrape_jobs[suburb_id] = {
            'status': 'error',
            'progress': f'Error: {str(e)}',
            'error': str(e),
        }


if __name__ == '__main__':
    init_db()
    # Auto-backup on every startup
    backup_db()
    logger.info("Database backed up on startup")
    cleaned = cleanup_agent_entries()
    if cleaned:
        logger.info(f"Cleaned up {cleaned} agent profile entries from DB")
    app.run(host='0.0.0.0', port=5000, debug=True)
