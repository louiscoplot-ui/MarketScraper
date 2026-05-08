"""Excel export endpoint — extracted from app.py to keep that module
under the MCP push size limit.

Builds a polished workbook (native Tables, currency/date formatting,
hyperlinks, status colors, stale DOM highlight) for the listings the
dashboard currently has selected. Wired via register_export_routes(app).
"""

import io
import re as _re
import logging
from datetime import datetime
from flask import request, send_file

from database import get_db, get_listings

logger = logging.getLogger(__name__)


def _calc_dom(listing):
    date_str = listing.get('listing_date') or ''
    if not date_str:
        return None
    ddmm = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    try:
        if ddmm:
            start = datetime(int(ddmm.group(3)), int(ddmm.group(2)), int(ddmm.group(1)))
        else:
            start = datetime.fromisoformat(date_str.replace('Z', ''))
    except (ValueError, TypeError):
        return None
    return max(0, (datetime.utcnow() - start).days)


def _parse_price_numeric(text):
    """Extract a plausible AUD value from REIWA's free-form price string."""
    if not text:
        return None
    s = str(text).replace(',', '').replace(' ', '')
    m = _re.search(r'\$?(\d+(?:\.\d+)?)\s*([MmKk]?)', s)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except ValueError:
        return None
    suffix = (m.group(2) or '').lower()
    if suffix == 'm':
        val *= 1_000_000
    elif suffix == 'k':
        val *= 1_000
    return int(val) if val >= 10_000 else None


def _parse_size_numeric(text):
    if not text:
        return None
    m = _re.search(r'(\d[\d,]*)', str(text))
    if not m:
        return None
    try:
        return int(m.group(1).replace(',', ''))
    except ValueError:
        return None


def _parse_listing_date(text):
    if not text:
        return None
    m = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', str(text).strip())
    if not m:
        return None
    try:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
    except ValueError:
        return None


def _iso_to_date(iso):
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso.replace('Z', '').split('.')[0]).date()
    except (ValueError, AttributeError):
        return None


def export_listings():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    STATUS_COLORS = {
        'active':      '15803D',
        'under_offer': 'B45309',
        'sold':        '4B5563',
        'withdrawn':   'B91C1C',
    }

    HEADER_FILL = PatternFill('solid', fgColor='1E3A2B')
    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BAND_FILL = PatternFill('solid', fgColor='F4F7F5')
    TOTAL_FILL = PatternFill('solid', fgColor='F5E6A8')
    TOTAL_FONT = Font(name='Calibri', bold=True, color='1E3A2B', size=11)
    BODY_FONT = Font(name='Calibri', size=10)
    BODY_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', indent=1)
    BODY_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    BODY_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', indent=1)
    SUBTLE_BORDER = Border(bottom=Side(style='thin', color='D4E0D8'))

    TABLE_STYLE = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                 showColumnStripes=False, showFirstColumn=False,
                                 showLastColumn=False)

    def _style_header_row(sheet, n_cols, height=30):
        sheet.row_dimensions[1].height = height
        for c in range(1, n_cols + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = SUBTLE_BORDER

    suburb_ids_str = request.args.get('suburb_ids', '')
    statuses_str = request.args.get('statuses', '')
    agent = request.args.get('agent', '').strip()
    agency = request.args.get('agency', '').strip()

    suburb_ids = None
    if suburb_ids_str:
        try:
            suburb_ids = [int(x) for x in suburb_ids_str.split(',') if x.strip()]
        except ValueError:
            pass
    statuses = None
    if statuses_str:
        statuses = [s.strip() for s in statuses_str.split(',') if s.strip()]

    # Per-user suburb scoping — same pattern as /api/listings (app.py:236).
    # Non-admin callers can't widen scope by passing arbitrary suburb_ids
    # in the query string. Empty intersection → empty workbook (200 with
    # header-only sheets — verified safe; no division-by-zero or empty-
    # list access in the downstream workbook construction).
    from admin_api import resolve_request_scope
    _, allowed = resolve_request_scope()
    if allowed is not None:
        if not allowed:
            listings = []
        else:
            if suburb_ids:
                suburb_ids = [s for s in suburb_ids if s in allowed]
                if not suburb_ids:
                    listings = []
                else:
                    listings = get_listings(suburb_ids=suburb_ids, statuses=statuses)
            else:
                listings = get_listings(suburb_ids=list(allowed), statuses=statuses)
    else:
        listings = get_listings(suburb_ids=suburb_ids, statuses=statuses)
    if agent:
        listings = [l for l in listings if l.get('agent') == agent]
    if agency:
        listings = [l for l in listings if l.get('agency') == agency]

    listings = sorted(
        listings,
        key=lambda l: _parse_listing_date(l.get('listing_date')) or datetime.min.date(),
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    columns = ['Address', 'Suburb', 'Price', 'Price (AUD)', 'Bed', 'Bath', 'Car',
               'Land (m²)', 'Internal (m²)', 'Agency', 'Agent', 'Listed', 'DOM',
               'Withdrawn', 'Status', 'Type', 'Link']
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    stale_font = Font(name='Calibri', bold=True, color='CC0000', size=10)
    col_align = {
        1: BODY_ALIGN_LEFT, 2: BODY_ALIGN_LEFT, 3: BODY_ALIGN_LEFT,
        4: BODY_ALIGN_RIGHT, 5: BODY_ALIGN_CENTER, 6: BODY_ALIGN_CENTER,
        7: BODY_ALIGN_CENTER, 8: BODY_ALIGN_RIGHT, 9: BODY_ALIGN_RIGHT,
        10: BODY_ALIGN_LEFT, 11: BODY_ALIGN_LEFT, 12: BODY_ALIGN_CENTER,
        13: BODY_ALIGN_CENTER, 14: BODY_ALIGN_CENTER, 15: BODY_ALIGN_CENTER,
        16: BODY_ALIGN_CENTER, 17: BODY_ALIGN_CENTER,
    }

    for row_idx, l in enumerate(listings, 2):
        status_raw = (l.get('status') or '').lower()
        price_text = l.get('price_text') or ''
        price_num = _parse_price_numeric(price_text)
        land_num = _parse_size_numeric(l.get('land_size'))
        internal_num = _parse_size_numeric(l.get('internal_size'))
        listed_date = _parse_listing_date(l.get('listing_date'))
        url = l.get('reiwa_url') or ''
        dom = _calc_dom(l)
        row_fill = BAND_FILL if row_idx % 2 == 0 else None

        ws.cell(row=row_idx, column=1, value=l.get('address', ''))
        ws.cell(row=row_idx, column=2, value=l.get('suburb_name', ''))
        ws.cell(row=row_idx, column=3, value=price_text)
        pcell = ws.cell(row=row_idx, column=4, value=price_num)
        if price_num is not None:
            pcell.number_format = '"$"#,##0'
        ws.cell(row=row_idx, column=5, value=l.get('bedrooms'))
        ws.cell(row=row_idx, column=6, value=l.get('bathrooms'))
        ws.cell(row=row_idx, column=7, value=l.get('parking'))
        lcell = ws.cell(row=row_idx, column=8, value=land_num)
        if land_num is not None:
            lcell.number_format = '#,##0" m²"'
        icell = ws.cell(row=row_idx, column=9, value=internal_num)
        if internal_num is not None:
            icell.number_format = '#,##0" m²"'
        ws.cell(row=row_idx, column=10, value=l.get('agency', ''))
        ws.cell(row=row_idx, column=11, value=l.get('agent', ''))
        dcell = ws.cell(row=row_idx, column=12, value=listed_date)
        if listed_date:
            dcell.number_format = 'DD/MM/YYYY'
        dom_cell = ws.cell(row=row_idx, column=13, value=dom)
        withdrawn_date = _iso_to_date(l.get('withdrawn_date'))
        wcell = ws.cell(row=row_idx, column=14, value=withdrawn_date)
        if withdrawn_date:
            wcell.number_format = 'DD/MM/YYYY'
        scell = ws.cell(row=row_idx, column=15,
                        value=status_raw.replace('_', ' ').title() if status_raw else '')
        ws.cell(row=row_idx, column=16, value=l.get('listing_type', ''))
        link_cell = ws.cell(row=row_idx, column=17, value="View on REIWA" if url else '')
        if url:
            link_cell.hyperlink = url

        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = BODY_FONT
            cell.alignment = col_align[c]
            cell.border = SUBTLE_BORDER
            if row_fill:
                cell.fill = row_fill

        if dom is not None and dom >= 60:
            dom_cell.font = stale_font
        color = STATUS_COLORS.get(status_raw)
        if color:
            scell.font = Font(name='Calibri', bold=True, color=color, size=10)
        if url:
            link_cell.font = Font(name='Calibri', color='1E3A8A', underline='single', size=10)

        ws.row_dimensions[row_idx].height = 22

    for col_idx, width in enumerate(
        [34, 18, 20, 14, 6, 6, 6, 13, 15, 28, 22, 12, 8, 12, 14, 14, 18], 1
    ):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    _style_header_row(ws, len(columns), height=34)
    ws.freeze_panes = "A2"

    last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
    last_row = max(2, len(listings) + 1)
    listings_table = Table(displayName="ListingsTable",
                           ref=f"A1:{last_col_letter}{last_row}")
    listings_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(listings_table)

    def _write_summary(sheet_name, table_name, headers, rows_data, col_widths):
        s = wb.create_sheet(sheet_name)
        for col_idx, h in enumerate(headers, 1):
            s.cell(row=1, column=col_idx, value=h)
        for row_idx, row_vals in enumerate(rows_data, 2):
            row_fill = BAND_FILL if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row_vals, 1):
                cell = s.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN_LEFT if col_idx == 1 else BODY_ALIGN_CENTER
                cell.border = SUBTLE_BORDER
                if row_fill:
                    cell.fill = row_fill
            s.row_dimensions[row_idx].height = 20
        if rows_data:
            totals_row = len(rows_data) + 2
            lbl = s.cell(row=totals_row, column=1, value='Total')
            lbl.font = TOTAL_FONT
            lbl.fill = TOTAL_FILL
            lbl.alignment = BODY_ALIGN_LEFT
            for col_idx in range(2, len(headers) + 1):
                letter = s.cell(row=1, column=col_idx).column_letter
                cell = s.cell(row=totals_row, column=col_idx,
                              value=f"=SUM({letter}2:{letter}{totals_row - 1})")
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.alignment = BODY_ALIGN_CENTER
            s.row_dimensions[totals_row].height = 22
        for col_idx, w in enumerate(col_widths, 1):
            s.column_dimensions[s.cell(row=1, column=col_idx).column_letter].width = w
        _style_header_row(s, len(headers), height=30)
        s.freeze_panes = "A2"
        last_data_row = max(2, len(rows_data) + 1)
        last_letter = s.cell(row=1, column=len(headers)).column_letter
        tbl = Table(displayName=table_name, ref=f"A1:{last_letter}{last_data_row}")
        tbl.tableStyleInfo = TABLE_STYLE
        s.add_table(tbl)

    def _group_stats(listings_, key):
        buckets = {}
        for l in listings_:
            k = l.get(key) or 'Unknown'
            b = buckets.setdefault(k, {'active': 0, 'under_offer': 0, 'sold': 0,
                                       'withdrawn': 0, 'total': 0})
            b['total'] += 1
            st = (l.get('status') or 'active').lower()
            if st in b:
                b[st] += 1
        rows = [
            [k, v['total'], v['active'], v['under_offer'], v['sold'], v['withdrawn']]
            for k, v in buckets.items()
        ]
        rows.sort(key=lambda r: -r[1])
        return rows

    _write_summary('Agents', 'AgentsTable',
                   ['Agent', 'Total', 'Active', 'Under Offer', 'Sold', 'Withdrawn'],
                   _group_stats(listings, 'agent'),
                   [26, 10, 10, 14, 10, 12])

    _write_summary('Agencies', 'AgenciesTable',
                   ['Agency', 'Total', 'Active', 'Under Offer', 'Sold', 'Withdrawn'],
                   _group_stats(listings, 'agency'),
                   [34, 10, 10, 14, 10, 12])

    suburb_names = set(l.get('suburb_name', '') for l in listings)
    if len(suburb_names) > 1:
        suburb_data = {}
        for l in listings:
            sn = l.get('suburb_name', 'Unknown')
            d = suburb_data.setdefault(sn, {
                'active': 0, 'under_offer': 0, 'sold': 0, 'withdrawn': 0,
                'agents': set(), 'agencies': set(),
            })
            st = (l.get('status') or 'active').lower()
            if st in d:
                d[st] += 1
            if l.get('agent'):
                d['agents'].add(l['agent'])
            if l.get('agency'):
                d['agencies'].add(l['agency'])
        suburb_rows = []
        for sname, data in sorted(suburb_data.items()):
            total = data['active'] + data['under_offer'] + data['sold'] + data['withdrawn']
            suburb_rows.append([
                sname, total, data['active'], data['under_offer'], data['sold'],
                data['withdrawn'], len(data['agents']), len(data['agencies'])
            ])
        _write_summary('Suburb Summary', 'SuburbsTable',
                       ['Suburb', 'Total', 'Active', 'Under Offer', 'Sold',
                        'Withdrawn', 'Agents', 'Agencies'],
                       suburb_rows,
                       [22, 10, 10, 14, 10, 12, 10, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = datetime.now().strftime('%Y-%m-%d')
    suburb_label = ''
    if suburb_ids:
        conn = get_db()
        names = []
        for sid in suburb_ids:
            row = conn.execute("SELECT name FROM suburbs WHERE id = ?", (sid,)).fetchone()
            if row:
                names.append(row['name'])
        conn.close()
        if len(names) <= 3:
            suburb_label = '_' + '_'.join(names)
        else:
            suburb_label = f'_{len(names)}_suburbs'

    filename = f'SuburbDesk{suburb_label}_{date_str}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def register_export_routes(app):
    app.add_url_rule(
        '/api/listings/export',
        endpoint='export_listings',
        view_func=export_listings,
        methods=['GET'],
    )
    logger.info("Export routes registered: /api/listings/export")
