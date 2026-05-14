"""Rental module API — separate from sales pipeline.

Wired into app.py via register_rental_routes(app). All rental data
lives in rental_listings / rental_owners / rental_suburbs; the users
table has a per-user rental_access flag gating these endpoints.

Routes:
  GET    /api/rentals/suburbs                       — list user's allowed rental suburbs
  GET    /api/rentals/<suburb>                      — listings + joined owners
  PATCH  /api/rentals/owner                         — UPSERT operator notes
  POST   /api/rentals/import                        — multipart .xlsx → bulk merge
  GET    /api/admin/rental-suburbs                  — admin allowlist read
  POST   /api/admin/rental-suburbs                  — admin add
  PATCH  /api/admin/rental-suburbs/<id>             — admin toggle/rename
  DELETE /api/admin/rental-suburbs/<id>             — admin cascade delete
  PATCH  /api/admin/users/<id>/rental-access        — admin toggle flag
"""

import io
import logging
from datetime import datetime
from flask import request, jsonify

from database import get_db, get_db_conn, USE_POSTGRES
from admin_api import (
    get_current_user, _require_admin,
    get_user_allowed_suburb_names,
)


logger = logging.getLogger(__name__)


# 10 MB upload cap — same ballpark as the RP Data import. Larger files
# almost certainly indicate the wrong template was uploaded.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# Excel header → DB column. Matches the spec's column list exactly.
# Lowercased for case-insensitive lookup.
_EXCEL_COLUMNS = {
    'status': 'status',
    'date listed': 'date_listed',
    'days on market': 'days_on_market',
    'date leased': 'date_leased',
    'address': 'address',
    'suburb': 'suburb',
    'price/week': 'price_week',
    'type': 'property_type',
    'bedrooms': 'beds',
    'bathrooms': 'baths',
    'parking': 'cars',
    'agency': 'agency',
    'agent': 'agent',
    'owner name': 'owner_name',
    'owner phone': 'owner_phone',
    'notes': 'notes',
    'link': 'url',
}


def _user_has_rental_access(user):
    """Admin → True. Non-admin → True only if users.rental_access is
    truthy (DB stores 0/1 in SQLite, bool in psycopg2)."""
    if not user:
        return False
    if (user.get('role') or '').lower() == 'admin':
        return True
    return bool(user.get('rental_access'))


def _resolve_rental_scope():
    """Three-state return:
      (user, None)  → rental-eligible (admin OR users.rental_access=1)
                      with full access to every active rental_suburb.
      (user, False) → authenticated but no rental_access flag — caller
                      should return 403.
      (None, None)  → unauthenticated.

    Rental coverage NO LONGER intersects with the sales user_suburbs
    table — there's no rental_user_suburbs allowlist yet, so any user
    granted rental_access sees every active rental_suburb. When per-
    user scoping is needed, add a rental_user_suburbs join table and
    branch here."""
    user = get_current_user()
    if not user:
        return None, None
    if (user.get('role') or '').lower() == 'admin':
        return user, None
    if not _user_has_rental_access(user):
        return user, False
    return user, None


def _allowed_rental_suburb_rows(conn, scope_names):
    """Returns every active rental_suburb as a list of dicts {id, name,
    active}. `scope_names` is kept in the signature for forward-compat
    when per-user rental scoping lands — today it's effectively
    ignored (None for all eligible users)."""
    rows = conn.execute(
        "SELECT id, name, active FROM rental_suburbs WHERE active = 1 ORDER BY name"
    ).fetchall()
    return [dict(r) for r in rows]


def register_rental_routes(app):

    # ------------------------------------------------------------------
    # GET /api/rentals/suburbs — user-facing dropdown source
    # ------------------------------------------------------------------
    @app.route('/api/rentals/suburbs', methods=['GET'])
    def rentals_list_suburbs():
        user, scope = _resolve_rental_scope()
        if not user:
            return jsonify({'error': 'Unauthenticated — provide X-Access-Key'}), 401
        if scope is False:
            return jsonify({'error': 'Rental access not granted'}), 403
        with get_db_conn() as conn:
            rows = _allowed_rental_suburb_rows(conn, scope)
        return jsonify({
            'suburbs': [{'id': r['id'], 'name': r['name']} for r in rows]
        })

    # ------------------------------------------------------------------
    # GET /api/rentals/export — multi-sheet Excel download
    # Registered BEFORE the /<path:suburb> route so Flask doesn't
    # route "/api/rentals/export" into the per-suburb handler with
    # suburb='export'.
    # ------------------------------------------------------------------
    @app.route('/api/rentals/export', methods=['GET'])
    def rentals_export_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        from datetime import date as _date

        user, scope = _resolve_rental_scope()
        if not user:
            return jsonify({'error': 'Unauthenticated — provide X-Access-Key'}), 401
        if scope is False:
            return jsonify({'error': 'Rental access not granted'}), 403

        single_suburb = (request.args.get('suburb') or '').strip()

        # Style palette — same family as export_api.py so the workbook
        # feels like part of the same product. Status colours diverge
        # from sales (sales uses green/amber/grey/red on listing
        # statuses; rental uses blue/teal/slate to match RentalView).
        HEADER_FILL = PatternFill('solid', fgColor='1E293B')
        HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        BAND_FILL = PatternFill('solid', fgColor='F8FAFC')
        BODY_FONT = Font(name='Calibri', size=10)
        BODY_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', indent=1)
        BODY_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
        SUBTLE_BORDER = Border(bottom=Side(style='thin', color='E2E8F0'))
        STATUS_STYLE = {
            'New':    (PatternFill('solid', fgColor='EFF6FF'),
                       Font(name='Calibri', size=10, color='1E40AF', bold=True)),
            'Active': (PatternFill('solid', fgColor='F0FDFA'),
                       Font(name='Calibri', size=10, color='0F766E', bold=True)),
            'Leased': (PatternFill('solid', fgColor='F8FAFC'),
                       Font(name='Calibri', size=10, color='64748B', italic=True)),
        }

        COLS = [
            ('status',         'Status',       12),
            ('address',        'Address',      32),
            ('suburb',         'Suburb',       18),
            ('price_week',     'Price/Week',   12),
            ('property_type',  'Type',         12),
            ('beds',           'Beds',         6),
            ('baths',          'Baths',        6),
            ('cars',           'Cars',         6),
            ('agency',         'Agency',       22),
            ('agent',          'Agent',        20),
            ('date_listed',    'Date Listed',  13),
            ('days_on_market', 'DOM',          7),
            ('date_leased',    'Date Leased',  13),
            ('owner_name',     'Owner Name',   22),
            ('owner_phone',    'Owner Phone',  16),
            ('notes',          'Notes',        34),
            ('url',            'Link',         16),
        ]

        conn = get_db()
        try:
            # Resolve the list of suburbs to include.
            if single_suburb:
                rows = conn.execute(
                    "SELECT name FROM rental_suburbs WHERE active = 1 "
                    "AND LOWER(name) = LOWER(?)",
                    (single_suburb,)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT name FROM rental_suburbs WHERE active = 1 ORDER BY name"
                ).fetchall()
            suburb_names = [dict(r)['name'] for r in rows]
            if not suburb_names:
                return jsonify({'error': 'No matching active rental suburbs.'}), 404

            # Pull each suburb's joined data — one query per sheet so
            # the workbook can stream large suburbs without holding
            # everything in memory at once. Same JOIN as the table
            # endpoint, identical ORDER BY.
            wb = Workbook()
            wb.remove(wb.active)

            # SUMMARY sheet — running tallies. Always present so the
            # operator can sanity-check counts before reading the
            # detail tabs.
            summary = wb.create_sheet('SUMMARY')
            summary_headers = ['Suburb', 'Total', 'New', 'Active', 'Leased']
            for ci, h in enumerate(summary_headers, 1):
                c = summary.cell(row=1, column=ci, value=h)
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = HEADER_ALIGN
            summary.row_dimensions[1].height = 24
            for ci, w in enumerate([22, 10, 10, 10, 10], 1):
                summary.column_dimensions[chr(64 + ci)].width = w

            sum_row = 2
            grand = {'total': 0, 'New': 0, 'Active': 0, 'Leased': 0}

            for name in suburb_names:
                listings = conn.execute(
                    """
                    SELECT
                      l.address, l.suburb, l.status, l.price_week,
                      l.property_type, l.beds, l.baths, l.cars,
                      l.agency, l.agent, l.date_listed, l.days_on_market,
                      l.date_leased, l.url,
                      COALESCE(o.owner_name, '')  AS owner_name,
                      COALESCE(o.owner_phone, '') AS owner_phone,
                      COALESCE(o.notes, '')       AS notes
                    FROM rental_listings l
                    LEFT JOIN rental_owners o
                      ON o.address = l.address AND o.suburb = l.suburb
                    WHERE LOWER(l.suburb) = LOWER(?)
                    ORDER BY
                      CASE l.status WHEN 'New' THEN 0 WHEN 'Active' THEN 1 ELSE 2 END,
                      l.date_listed DESC
                    """,
                    (name,)
                ).fetchall()
                listings = [dict(r) for r in listings]

                # Summary tally for this suburb.
                per_status = {'New': 0, 'Active': 0, 'Leased': 0}
                for r in listings:
                    s = (r.get('status') or '').strip()
                    if s in per_status:
                        per_status[s] += 1
                total = len(listings)
                summary.cell(row=sum_row, column=1, value=name).alignment = BODY_ALIGN_LEFT
                summary.cell(row=sum_row, column=2, value=total).alignment = BODY_ALIGN_CENTER
                summary.cell(row=sum_row, column=3, value=per_status['New']).alignment = BODY_ALIGN_CENTER
                summary.cell(row=sum_row, column=4, value=per_status['Active']).alignment = BODY_ALIGN_CENTER
                summary.cell(row=sum_row, column=5, value=per_status['Leased']).alignment = BODY_ALIGN_CENTER
                for ci in range(1, 6):
                    summary.cell(row=sum_row, column=ci).font = BODY_FONT
                if sum_row % 2 == 0:
                    for ci in range(1, 6):
                        summary.cell(row=sum_row, column=ci).fill = BAND_FILL
                sum_row += 1
                grand['total'] += total
                for k in ('New', 'Active', 'Leased'):
                    grand[k] += per_status[k]

                # Suburb sheet. Excel limits sheet names to 31 chars
                # and forbids `/ : * ? [ ] \`; rental suburb names are
                # short clean strings so a simple slice is enough.
                safe_sheet = name[:31].replace('/', '_').replace('\\', '_')
                ws = wb.create_sheet(safe_sheet)
                for ci, (_k, label, width) in enumerate(COLS, 1):
                    c = ws.cell(row=1, column=ci, value=label)
                    c.fill = HEADER_FILL
                    c.font = HEADER_FONT
                    c.alignment = HEADER_ALIGN
                    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width
                ws.row_dimensions[1].height = 26
                ws.freeze_panes = 'A2'

                for ri, rec in enumerate(listings, start=2):
                    status = (rec.get('status') or '').strip()
                    status_fill, status_font = STATUS_STYLE.get(status, (None, None))
                    for ci, (key, _label, _w) in enumerate(COLS, 1):
                        val = rec.get(key)
                        if val is None:
                            val = ''
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.font = BODY_FONT
                        c.border = SUBTLE_BORDER
                        if key == 'status' and status_fill is not None:
                            c.fill = status_fill
                            c.font = status_font
                            c.alignment = BODY_ALIGN_CENTER
                        elif key in ('beds', 'baths', 'cars', 'days_on_market'):
                            c.alignment = BODY_ALIGN_CENTER
                            if ri % 2 == 0:
                                c.fill = BAND_FILL
                        else:
                            c.alignment = BODY_ALIGN_LEFT
                            if ri % 2 == 0:
                                c.fill = BAND_FILL
        finally:
            conn.close()

        # Grand-total row on SUMMARY — yellow band so it stands out.
        TOTAL_FILL = PatternFill('solid', fgColor='FEF3C7')
        TOTAL_FONT = Font(name='Calibri', bold=True, color='1E293B', size=11)
        summary.cell(row=sum_row, column=1, value='TOTAL').font = TOTAL_FONT
        summary.cell(row=sum_row, column=2, value=grand['total']).font = TOTAL_FONT
        summary.cell(row=sum_row, column=3, value=grand['New']).font = TOTAL_FONT
        summary.cell(row=sum_row, column=4, value=grand['Active']).font = TOTAL_FONT
        summary.cell(row=sum_row, column=5, value=grand['Leased']).font = TOTAL_FONT
        for ci in range(1, 6):
            summary.cell(row=sum_row, column=ci).fill = TOTAL_FILL
            summary.cell(row=sum_row, column=ci).alignment = (
                BODY_ALIGN_LEFT if ci == 1 else BODY_ALIGN_CENTER
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = _date.today().strftime('%Y-%m-%d')
        if single_suburb:
            safe = single_suburb.replace(' ', '_').replace('/', '_')
            filename = f'rental_export_{safe}_{date_str}.xlsx'
        else:
            filename = f'rental_export_{date_str}.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    # ------------------------------------------------------------------
    # GET /api/rentals/<suburb> — table content
    # ------------------------------------------------------------------
    @app.route('/api/rentals/<path:suburb>', methods=['GET'])
    def rentals_get_suburb(suburb):
        user, scope = _resolve_rental_scope()
        if not user:
            return jsonify({'error': 'Unauthenticated — provide X-Access-Key'}), 401
        if scope is False:
            return jsonify({'error': 'Rental access not granted'}), 403
        suburb_clean = (suburb or '').strip()
        if not suburb_clean:
            return jsonify({'error': 'suburb required'}), 400
        if scope is not None and suburb_clean.lower() not in scope:
            return jsonify({'error': 'Not authorised for that suburb'}), 403
        # Status sort: New (0), Active (1), Leased (2) — ORDER BY CASE.
        # Then date_listed DESC so the freshest within each bucket lands
        # at the top. Coalesce owner fields from rental_owners so the UI
        # can render every column without a second round-trip.
        # SELECT only the columns the UI actually renders — first_seen
        # / last_seen / id are unused on the rental page and trimming
        # them shaves a few KB off the payload on busy suburbs.
        with get_db_conn() as conn:
            rows = conn.execute(
                """
                SELECT
                  l.address, l.suburb, l.status, l.price_week,
                  l.property_type, l.beds, l.baths, l.cars,
                  l.agency, l.agent, l.date_listed, l.days_on_market,
                  l.date_leased, l.url,
                  COALESCE(o.owner_name, '')  AS owner_name,
                  COALESCE(o.owner_phone, '') AS owner_phone,
                  COALESCE(o.notes, '')       AS notes
                FROM rental_listings l
                LEFT JOIN rental_owners o
                  ON o.address = l.address AND o.suburb = l.suburb
                WHERE LOWER(l.suburb) = LOWER(?)
                ORDER BY
                  CASE l.status WHEN 'New' THEN 0 WHEN 'Active' THEN 1 ELSE 2 END,
                  l.date_listed DESC
                """,
                (suburb_clean,)
            ).fetchall()
        return jsonify({
            'suburb': suburb_clean,
            'count': len(rows),
            'listings': [dict(r) for r in rows],
        })

    # ------------------------------------------------------------------
    # PATCH /api/rentals/owner — UPSERT operator data on rental_owners
    # ------------------------------------------------------------------
    @app.route('/api/rentals/owner', methods=['PATCH'])
    def rentals_patch_owner():
        user, scope = _resolve_rental_scope()
        if not user:
            return jsonify({'error': 'Unauthenticated — provide X-Access-Key'}), 401
        if scope is False:
            return jsonify({'error': 'Rental access not granted'}), 403
        body = request.get_json(silent=True) or {}
        address = (body.get('address') or '').strip()
        suburb = (body.get('suburb') or '').strip()
        if not address or not suburb:
            return jsonify({'error': 'address and suburb required'}), 400
        if scope is not None and suburb.lower() not in scope:
            return jsonify({'error': 'Not authorised for that suburb'}), 403

        # Partial-body contract: only the field KEYS present in the
        # request are written. Without this guard, the frontend's
        # race-prone "send-all-3-fields-every-blur" pattern wiped
        # sibling fields with a stale snapshot. We now read each
        # column's intent from key presence — `notes: ''` is a real
        # "clear notes" instruction, while omitting `notes` means
        # "leave notes alone".
        updates = {}
        for field in ('owner_name', 'owner_phone', 'notes'):
            if field in body:
                v = body[field]
                updates[field] = (v or '').strip() if isinstance(v, str) else ''
        if not updates:
            return jsonify({'error': 'No fields to update'}), 400

        conn = get_db()
        try:
            existing = conn.execute(
                "SELECT id FROM rental_owners WHERE address = ? AND suburb = ?",
                (address, suburb)
            ).fetchone()
            if existing is None:
                # INSERT — defaults '' for any field the body didn't carry.
                conn.execute(
                    "INSERT INTO rental_owners "
                    "(address, suburb, owner_name, owner_phone, notes) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (
                        address, suburb,
                        updates.get('owner_name', ''),
                        updates.get('owner_phone', ''),
                        updates.get('notes', ''),
                    )
                )
            else:
                # UPDATE — dynamic SET on only the fields actually sent.
                sets = [f"{k} = ?" for k in updates.keys()]
                params = list(updates.values())
                if USE_POSTGRES:
                    sets.append("updated_at = CURRENT_TIMESTAMP")
                else:
                    sets.append("updated_at = datetime('now')")
                params.extend([address, suburb])
                conn.execute(
                    f"UPDATE rental_owners SET {', '.join(sets)} "
                    "WHERE address = ? AND suburb = ?",
                    params
                )
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success': True, 'updated_fields': list(updates.keys())})

    # ------------------------------------------------------------------
    # POST /api/rentals/import — multipart Excel bulk merge
    # ------------------------------------------------------------------
    @app.route('/api/rentals/import', methods=['POST'])
    def rentals_import_excel():
        user, scope = _resolve_rental_scope()
        if not user:
            return jsonify({'error': 'Unauthenticated — provide X-Access-Key'}), 401
        if scope is False:
            return jsonify({'error': 'Rental access not granted'}), 403
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded. Use multipart field "file".'}), 400
        f = request.files['file']
        name = (f.filename or '').lower()
        if not (name.endswith('.xlsx') or name.endswith('.xls')):
            return jsonify({'error': 'Only .xlsx / .xls accepted'}), 400
        raw = f.read()
        if not raw:
            return jsonify({'error': 'Empty file'}), 400
        if len(raw) > MAX_UPLOAD_BYTES:
            return jsonify({'error': f'File too large (>{MAX_UPLOAD_BYTES // (1024 * 1024)} MB)'}), 400
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({'error': 'openpyxl not installed on the server'}), 500
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as e:
            return jsonify({'error': f'Could not parse workbook: {e}'}), 400

        # Non-destructive merge counters surfaced in the toast:
        #   inserted = brand-new rental_listings rows
        #   enriched = existing rows where AT LEAST ONE empty field
        #              (listing or owner) got filled by this import
        #   skipped  = invalid row (no address) OR fully-populated row
        #              that this Excel had nothing new to add to.
        # Listing columns that can be filled (status / price_week etc.).
        # `status` is included so a 'Leased' row from the Excel can
        # populate an existing 'Active' row IF the existing row's
        # status field is empty — which it shouldn't be in practice
        # (the scraper writes one of New/Active/Leased on every row),
        # so this is effectively a no-op safety net.
        FILLABLE = (
            'status', 'price_week', 'property_type', 'beds', 'baths',
            'cars', 'agency', 'agent', 'date_listed', 'days_on_market',
            'date_leased', 'url',
        )

        inserted = 0
        enriched = 0
        skipped = 0
        suburbs_seen = []
        conn = get_db()
        try:
            for sheet_name in wb.sheetnames:
                if sheet_name.strip().upper() == 'SUMMARY':
                    continue
                ws = wb[sheet_name]
                # Sheet name is the suburb. Honour scope — silently skip
                # sheets outside the caller's allowed suburbs.
                if scope is not None and sheet_name.strip().lower() not in scope:
                    continue

                header_idx = None
                col_map = {}
                # Scan first 5 rows for a header containing 'address'.
                rows_iter = list(ws.iter_rows(values_only=True))
                for i, row in enumerate(rows_iter[:5]):
                    cells = [('' if c is None else str(c)).strip().lower() for c in row]
                    if 'address' in cells and 'suburb' in cells:
                        header_idx = i
                        for j, cell in enumerate(cells):
                            if cell in _EXCEL_COLUMNS:
                                col_map[_EXCEL_COLUMNS[cell]] = j
                        break
                if header_idx is None:
                    # Sheet has no usable header — count once as skipped
                    # so the operator sees a non-zero number explaining
                    # the missing rows.
                    skipped += 1
                    continue
                suburbs_seen.append(sheet_name)

                for row in rows_iter[header_idx + 1:]:
                    def cell(key):
                        idx = col_map.get(key)
                        if idx is None or idx >= len(row):
                            return ''
                        v = row[idx]
                        return '' if v is None else str(v).strip()

                    addr = cell('address')
                    sub = cell('suburb') or sheet_name.strip()
                    if not addr or not sub:
                        skipped += 1
                        continue

                    # Snapshot of the existing listing — drives the
                    # field-level "fill empty only" merge below. None
                    # means we'll INSERT.
                    existing = conn.execute(
                        "SELECT status, price_week, property_type, beds, baths, "
                        "       cars, agency, agent, date_listed, days_on_market, "
                        "       date_leased, url FROM rental_listings "
                        "WHERE address = ? AND suburb = ?",
                        (addr, sub)
                    ).fetchone()

                    row_inserted = False
                    row_enriched = False

                    if existing is None:
                        # Brand-new row — full INSERT. status defaults to
                        # 'Active' when the Excel didn't carry one so the
                        # NOT NULL constraint stays satisfied.
                        conn.execute(
                            "INSERT INTO rental_listings "
                            "(address, suburb, status, price_week, property_type, "
                            " beds, baths, cars, agency, agent, date_listed, "
                            " days_on_market, date_leased, url) "
                            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (
                                addr, sub,
                                cell('status') or 'Active',
                                cell('price_week'), cell('property_type'),
                                cell('beds'), cell('baths'), cell('cars'),
                                cell('agency'), cell('agent'),
                                cell('date_listed'), cell('days_on_market'),
                                cell('date_leased'), cell('url'),
                            )
                        )
                        row_inserted = True
                    else:
                        # Field-level merge — only fill columns the DB
                        # has as NULL or empty. Never overwrite a
                        # populated cell. status is in FILLABLE but
                        # rental_listings.status defaults 'Active', so
                        # the DB value is virtually always non-empty
                        # and the status column stays untouched here.
                        existing_d = dict(existing)
                        sets = []
                        params = []
                        for field in FILLABLE:
                            db_val = (existing_d.get(field) or '').strip()
                            if db_val:
                                continue
                            excel_val = cell(field)
                            if not excel_val:
                                continue
                            sets.append(f"{field} = ?")
                            params.append(excel_val)
                        if sets:
                            # last_seen bumps because we touched the row,
                            # so the listings UI sorts the freshly-merged
                            # rows toward the recent end.
                            sets.append("last_seen = ?")
                            params.append(datetime.utcnow().isoformat())
                            params.extend([addr, sub])
                            conn.execute(
                                f"UPDATE rental_listings SET {', '.join(sets)} "
                                "WHERE address = ? AND suburb = ?",
                                params
                            )
                            row_enriched = True

                    # rental_owners — same fill-empty-only contract,
                    # field by field. Operator-typed values in the UI
                    # are sacred; the Excel only writes into gaps.
                    owner_name = cell('owner_name')
                    owner_phone = cell('owner_phone')
                    notes_text = cell('notes')
                    if owner_name or owner_phone or notes_text:
                        o_row = conn.execute(
                            "SELECT owner_name, owner_phone, notes FROM rental_owners "
                            "WHERE address = ? AND suburb = ?",
                            (addr, sub)
                        ).fetchone()
                        if o_row is None:
                            conn.execute(
                                "INSERT INTO rental_owners "
                                "(address, suburb, owner_name, owner_phone, notes) "
                                "VALUES (?, ?, ?, ?, ?)",
                                (addr, sub, owner_name, owner_phone, notes_text)
                            )
                            row_enriched = True
                        else:
                            od = dict(o_row)
                            o_sets = []
                            o_params = []
                            for field, val in (
                                ('owner_name',  owner_name),
                                ('owner_phone', owner_phone),
                                ('notes',       notes_text),
                            ):
                                if not val:
                                    continue
                                if (od.get(field) or '').strip():
                                    continue
                                o_sets.append(f"{field} = ?")
                                o_params.append(val)
                            if o_sets:
                                if USE_POSTGRES:
                                    o_sets.append("updated_at = CURRENT_TIMESTAMP")
                                else:
                                    o_sets.append("updated_at = datetime('now')")
                                o_params.extend([addr, sub])
                                conn.execute(
                                    f"UPDATE rental_owners SET {', '.join(o_sets)} "
                                    "WHERE address = ? AND suburb = ?",
                                    o_params
                                )
                                row_enriched = True

                    if row_inserted:
                        inserted += 1
                    elif row_enriched:
                        enriched += 1
                    else:
                        skipped += 1
            conn.commit()
        except Exception as e:
            logger.exception("Rental import failed")
            return jsonify({'error': f'Import crashed: {e}'}), 500
        finally:
            conn.close()
        return jsonify({
            'inserted': inserted,
            'enriched': enriched,
            'skipped': skipped,
            'suburbs': suburbs_seen,
        })

    # ------------------------------------------------------------------
    # Admin allowlist CRUD
    # ------------------------------------------------------------------
    @app.route('/api/admin/rental-suburbs', methods=['GET'])
    def admin_list_rental_suburbs():
        _u, err = _require_admin()
        if err:
            return err
        conn = get_db()
        rows = conn.execute(
            "SELECT id, name, active, created_at FROM rental_suburbs ORDER BY name"
        ).fetchall()
        conn.close()
        return jsonify({'suburbs': [dict(r) for r in rows]})

    @app.route('/api/admin/rental-suburbs', methods=['POST'])
    def admin_add_rental_suburb():
        _u, err = _require_admin()
        if err:
            return err
        body = request.get_json(silent=True) or {}
        name = (body.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'name required'}), 400
        conn = get_db()
        try:
            conn.execute(
                "INSERT INTO rental_suburbs (name) VALUES (?)", (name,)
            )
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Could not add: {e}'}), 400
        row = conn.execute(
            "SELECT id, name, active FROM rental_suburbs WHERE name = ?", (name,)
        ).fetchone()
        conn.close()
        return jsonify(dict(row) if row else {'name': name})

    @app.route('/api/admin/rental-suburbs/<int:sid>', methods=['PATCH'])
    def admin_patch_rental_suburb(sid):
        _u, err = _require_admin()
        if err:
            return err
        body = request.get_json(silent=True) or {}
        sets = []
        params = []
        if 'active' in body:
            sets.append("active = ?")
            params.append(1 if body['active'] else 0)
        if 'name' in body:
            new_name = (body.get('name') or '').strip()
            if not new_name:
                return jsonify({'error': 'name cannot be empty'}), 400
            sets.append("name = ?")
            params.append(new_name)
        if not sets:
            return jsonify({'error': 'No fields to update'}), 400
        params.append(sid)
        conn = get_db()
        try:
            conn.execute(
                f"UPDATE rental_suburbs SET {', '.join(sets)} WHERE id = ?", params
            )
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Update failed: {e}'}), 400
        row = conn.execute(
            "SELECT id, name, active FROM rental_suburbs WHERE id = ?", (sid,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        return jsonify(dict(row))

    @app.route('/api/admin/rental-suburbs/batch', methods=['PATCH'])
    def admin_batch_patch_rental_suburbs():
        """Apply many active/inactive flips in one DB round-trip — the
        admin panel uses this instead of N individual PATCH calls when
        the operator clicks Save after multi-select edits. Body shape:
            { "updates": [ { "id": int, "active": bool }, ... ] }
        Unknown ids silently dropped; partial success returns the count
        of rows actually touched (rowcount may differ from len(updates)
        when some ids no longer exist)."""
        _u, err = _require_admin()
        if err:
            return err
        body = request.get_json(silent=True) or {}
        updates = body.get('updates') or []
        if not isinstance(updates, list):
            return jsonify({'error': 'updates must be a list'}), 400
        clean = []
        for u in updates:
            if not isinstance(u, dict):
                continue
            try:
                sid = int(u.get('id'))
            except (TypeError, ValueError):
                continue
            active_flag = 1 if u.get('active') else 0
            clean.append((active_flag, sid))
        if not clean:
            return jsonify({'updated': 0})
        conn = get_db()
        updated = 0
        try:
            for active_flag, sid in clean:
                cur = conn.execute(
                    "UPDATE rental_suburbs SET active = ? WHERE id = ?",
                    (active_flag, sid)
                )
                if cur.rowcount and cur.rowcount > 0:
                    updated += cur.rowcount
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Batch update failed: {e}'}), 500
        conn.close()
        return jsonify({'updated': updated})

    @app.route('/api/admin/rental-suburbs/<int:sid>', methods=['DELETE'])
    def admin_delete_rental_suburb(sid):
        _u, err = _require_admin()
        if err:
            return err
        conn = get_db()
        row = conn.execute(
            "SELECT name FROM rental_suburbs WHERE id = ?", (sid,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Not found'}), 404
        name = dict(row)['name']
        # Manual cascade — rental_listings / rental_owners reference
        # suburb by string, not FK, so DROP doesn't sweep them.
        conn.execute(
            "DELETE FROM rental_listings WHERE LOWER(suburb) = LOWER(?)", (name,)
        )
        conn.execute(
            "DELETE FROM rental_owners WHERE LOWER(suburb) = LOWER(?)", (name,)
        )
        conn.execute("DELETE FROM rental_suburbs WHERE id = ?", (sid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'deleted_suburb': name})

    @app.route('/api/admin/users/<int:user_id>/rental-access', methods=['PATCH'])
    def admin_patch_user_rental_access(user_id):
        _u, err = _require_admin()
        if err:
            return err
        body = request.get_json(silent=True) or {}
        if 'rental_access' not in body:
            return jsonify({'error': 'rental_access required'}), 400
        flag = 1 if body['rental_access'] else 0
        conn = get_db()
        row = conn.execute(
            "SELECT id FROM users WHERE id = ?", (user_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        conn.execute(
            "UPDATE users SET rental_access = ? WHERE id = ?", (flag, user_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'user_id': user_id, 'rental_access': bool(flag)})

    logger.info(
        "Rental routes: /api/rentals/{suburbs,<suburb>,owner,import} + "
        "/api/admin/rental-suburbs[/<id>] + "
        "/api/admin/users/<id>/rental-access"
    )
