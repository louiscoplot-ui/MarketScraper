"""Hot Vendor Lead Scoring — Excel report generator.

Builds a polished .xlsx from a `score_csv` result dict. Four sheets:
  1. Summary            (suburb profile, auto-calibrated weights, score distribution)
  2. All Scored         (full property list with all per-factor scores)
  3. High Probability   (HOT + WARM filtered + sorted)
  4. Houses Priority    (Houses only, with full details)

Mail-ready and consistent with the v4 scoring methodology.
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Palette
NAVY = '1B2A4A'
WHITE = 'FFFFFF'
MID = '2E6DA4'
ALT = 'EBF5FB'
HRL = 'FADBD8'   # HOT pale red
WOL = 'FDEBD0'   # WARM pale orange
MYL = 'FEF9E7'   # MEDIUM pale yellow
LGL = 'F2F3F4'   # LOW pale grey
HR_TXT = 'C0392B'
WO_TXT = 'E67E22'
MY_TXT = '7D6608'
LGY_TXT = '95A5A6'

CAT_FILL = {'HOT': HRL, 'WARM': WOL, 'MEDIUM': MYL, 'LOW': LGL}
CAT_TEXT = {'HOT': HR_TXT, 'WARM': WO_TXT, 'MEDIUM': MY_TXT, 'LOW': LGY_TXT}


def _font(sz=10, bold=False, color=NAVY):
    return Font(name='Arial', size=sz, bold=bold, color=color)


def _fill(color=WHITE):
    return PatternFill('solid', fgColor=color)


def _ctr():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _lft(indent=0):
    return Alignment(horizontal='left', vertical='center', indent=indent, wrap_text=True)


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws, row, text, ncols, bg=NAVY, fg=WHITE, sz=12, h=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(bg)
    c.font = _font(sz, bold=True, color=fg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2, wrap_text=True)
    ws.row_dimensions[row].height = h


def _sub_row(ws, row, text, ncols, bg=MID, fg=WHITE, h=18):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(bg)
    c.font = _font(10, color=fg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[row].height = h


def _header_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.fill = _fill(NAVY)
        c.font = _font(10, bold=True, color=WHITE)
        c.alignment = _ctr()


def _fp(v):
    return '-' if v is None or v == '' else f'${int(v):,}'


def _fpct(v, d=1):
    if v is None or v == '' or (isinstance(v, float) and v != v):
        return '-'
    return f'{v:.{d}f}%'


def _fnum(v, d=1):
    if v is None or v == '' or (isinstance(v, float) and v != v):
        return '-'
    return f'{v:.{d}f}'


def _cs(v):
    if v is None or str(v) in ('-', 'nan', 'None', ''):
        return ''
    return str(v).strip()


def _build_summary(wb, result):
    ws = wb.create_sheet('1 - Summary')
    ws.sheet_view.showGridLines = False

    p = result['profile']
    w = result['weights']
    suburb = result['suburb']
    today = result['today']

    r = 1
    _title_row(ws, r, f"📊  HOT VENDOR SCORING — {suburb} — {today}", 4)
    r += 1
    _sub_row(ws, r, f"v4 Auto-Calibration  |  {result['kept_count']:,} properties retained "
                    f"({result['excluded_count']} excluded from {result['raw_count']:,})", 4)
    r += 2

    def block(title, rows):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1, title)
        c.fill = _fill('2C3E50')
        c.font = _font(10, bold=True, color=WHITE)
        c.alignment = _lft(1)
        ws.row_dimensions[r].height = 20
        r += 1
        for i, (k, v) in enumerate(rows):
            bg = ALT if i % 2 == 0 else WHITE
            kc = ws.cell(r, 1, k)
            vc = ws.cell(r, 2, v)
            kc.fill = vc.fill = _fill(bg)
            kc.font = _font(10, bold=True)
            vc.font = _font(10)
            kc.alignment = vc.alignment = _lft(1)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            r += 1
        r += 1

    block('Suburb Profile', [
        ('Classification',        f"{'Mature' if p['is_mature'] else 'Dynamic'} | "
                                  f"{'Premium' if p['is_premium'] else 'Standard'} | "
                                  f"{'High-gain' if p['is_high_gain'] else 'Moderate-gain'}"),
        ('Median holding',         f"{p['median_hold']} yrs"),
        ('% holding > 20 yrs',     f"{p['pct_long_hold']*100:.0f}%"),
        ('% with gain > 200%',     f"{p['pct_high_gain']*100:.0f}%"),
        ('% single-sale only',     f"{p['pct_1sale']*100:.0f}%"),
        ('Median sale price',      _fp(p['med_price'])),
        ('Median gain %',          _fpct(p['median_gain_pct'])),
    ])

    block('Auto-Calibrated Weights', [
        ('Holding Duration',       f"{w['hold']*100:.0f}%"),
        ('Property Type',          f"{w['type']*100:.0f}%"),
        ('Capital Gain %',         f"{w['gain']*100:.1f}%"),
        ('CAGR %/yr',              f"{w['cagr']*100:.1f}%"),
        ('Sale Frequency',         f"{w['freq']*100:.0f}%"),
        ('Estimated Latent Profit', f"{w['profit']*100:.0f}%"),
    ])

    rationale = ', '.join(result['rationale']) or 'standard'
    block('Calibration Rationale', [('Why these weights', rationale)])

    n_hot = sum(1 for x in result['properties'] if x['category'] == 'HOT')
    n_warm = sum(1 for x in result['properties'] if x['category'] == 'WARM')
    n_med = sum(1 for x in result['properties'] if x['category'] == 'MEDIUM')
    n_low = sum(1 for x in result['properties'] if x['category'] == 'LOW')
    total = max(len(result['properties']), 1)
    block('Score Distribution', [
        (f"🔴 HOT (≥{result['q_hot']:.1f})",       f"{n_hot:,} ({n_hot/total*100:.0f}%)"),
        (f"🟠 WARM ({result['q_warm']:.1f}–{result['q_hot']:.1f})", f"{n_warm:,} ({n_warm/total*100:.0f}%)"),
        (f"🟡 MEDIUM ({result['q_medium']:.1f}–{result['q_warm']:.1f})", f"{n_med:,} ({n_med/total*100:.0f}%)"),
        (f"⚪ LOW (<{result['q_medium']:.1f})",     f"{n_low:,} ({n_low/total*100:.0f}%)"),
        ('High-probability leads (HOT+WARM)', f"{n_hot+n_warm:,} properties"),
    ])

    block('Estimated Value', [
        ('Median $/m² Houses (recent)',     _fp(result['median_m2_house'])),
        ('Median $/m² Apartments (recent)', _fp(result['median_m2_apt'])),
    ])

    _set_widths(ws, [42, 20, 12, 12])
    ws.freeze_panes = 'A4'


def _write_property_row(ws, row, p, cols, bg_override=None):
    bg = bg_override or (ALT if row % 2 == 0 else WHITE)
    cat = p.get('category')
    for ci, (key, fmt) in enumerate(cols, 1):
        if key == 'category':
            v = f"🔴 {cat}" if cat == 'HOT' else f"🟠 {cat}" if cat == 'WARM' else f"🟡 {cat}" if cat == 'MEDIUM' else f"⚪ {cat}"
            c = ws.cell(row, ci, v)
            c.fill = _fill(CAT_FILL.get(cat, WHITE))
            c.font = _font(10, bold=cat in ('HOT', 'WARM'), color=CAT_TEXT.get(cat, NAVY))
            c.alignment = _ctr()
        else:
            raw = p.get(key)
            if fmt == 'price':
                v = _fp(raw)
            elif fmt == 'pct':
                v = _fpct(raw)
            elif fmt == 'pct2':
                v = _fpct(raw, 2)
            elif fmt == 'num':
                v = _fnum(raw)
            elif fmt == 'int':
                v = '-' if raw is None else int(raw)
            elif fmt == 'str':
                v = _cs(raw)
            elif fmt == 'date':
                v = _cs(raw)
            else:
                v = raw if raw is not None else '-'
            c = ws.cell(row, ci, v)
            c.fill = _fill(bg)
            c.font = _font(10)
            c.alignment = _ctr() if fmt in ('price', 'pct', 'pct2', 'num', 'int') else _lft(1)


SCORED_COLS = [
    ('rank', 'int'), ('address', 'str'), ('type', 'str'),
    ('bedrooms', 'int'), ('bathrooms', 'int'),
    ('last_sale_price', 'price'), ('owner_purchase_date', 'date'),
    ('holding_years', 'num'),
    ('owner_gain_dollars', 'int'), ('owner_gain_pct', 'pct'),
    ('cagr', 'pct2'), ('sales_count', 'int'),
    ('estimated_value', 'price'), ('potential_profit', 'price'),
    ('potential_profit_pct', 'pct'),
    ('hold_score', 'int'), ('type_score', 'int'), ('gain_score', 'int'),
    ('cagr_score', 'int'), ('freq_score', 'int'), ('prof_score', 'int'),
    ('final_score', 'num'), ('category', 'cat'),
]
SCORED_HEADERS = ['Rank', 'Address', 'Type', 'Bed', 'Bath', 'Last Sale $',
                  'Purchase Date', 'Holding (yrs)',
                  'Gain ($)', 'Gain %', 'CAGR %/yr',
                  '# Sales', 'Est. Value', 'Latent Profit $', 'Latent Profit %',
                  'Hold', 'Type', 'Gain', 'CAGR', 'Freq', 'Profit',
                  'Final Score', 'Category']
SCORED_WIDTHS = [6, 32, 10, 5, 5, 14, 16, 11, 12, 10, 10, 7, 14, 14, 12, 7, 7, 7, 7, 7, 7, 11, 11]


LEADS_COLS = SCORED_COLS + [
    ('current_owner', 'str'), ('agency', 'str'), ('agent', 'str'),
]
LEADS_HEADERS = SCORED_HEADERS + ['Current Owner', 'Agency (last sale)', 'Agent (last sale)']
LEADS_WIDTHS = SCORED_WIDTHS + [26, 26, 22]


def _build_data_sheet(wb, name, title, properties, columns, headers, widths):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    _title_row(ws, 1, title, len(headers))
    _sub_row(ws, 2, '⚠️  Verify on REIWA / agency CRM before prospecting — '
                    'Landgate registration lag up to 12 months.',
             len(headers), bg='FFF3CD', fg='856404')

    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 30
    _set_widths(ws, widths)
    ws.freeze_panes = 'A4'

    for i, p in enumerate(properties, 4):
        _write_property_row(ws, i, p, columns)

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f'A3:{last_col}{ws.max_row}'


def build_workbook(result):
    """Assemble the 4-sheet workbook in memory and return BytesIO."""
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary(wb, result)

    properties = result['properties']
    suburb = result['suburb']

    _build_data_sheet(
        wb, '2 - All Scored',
        f"📋  ALL SCORED PROPERTIES — {suburb} — {len(properties):,} entries",
        properties, SCORED_COLS, SCORED_HEADERS, SCORED_WIDTHS,
    )

    leads = [p for p in properties if p['category'] in ('HOT', 'WARM')]
    _build_data_sheet(
        wb, '3 - High Probability Leads',
        f"🎯  HIGH PROBABILITY LEADS — {suburb} — {len(leads):,} (HOT+WARM)",
        leads, LEADS_COLS, LEADS_HEADERS, LEADS_WIDTHS,
    )

    houses = [p for p in properties if p['type'] == 'House']
    _build_data_sheet(
        wb, '4 - Houses Priority',
        f"🏠  HOUSES PRIORITY — {suburb} — {len(houses):,} houses",
        houses, LEADS_COLS, LEADS_HEADERS, LEADS_WIDTHS,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def workbook_filename(suburb, today=None):
    today = today or date.today()
    safe = (suburb or 'Suburb').replace(' ', '_').replace('/', '_').title()
    return f"{safe}_HotVendor_v4_{today.strftime('%Y%m%d')}.xlsx"
