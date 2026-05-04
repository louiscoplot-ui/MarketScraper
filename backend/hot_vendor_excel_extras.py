"""Hot Vendor Excel — extra sheets that mirror the v4 standalone script.

Two heavy sheets split out of hot_vendor_excel.py to keep that module
small and pushable:
  - Methodology  : formula + per-factor scoring tables + auto-calibration rules
  - Market Analysis : suburb profile blocks, score distribution, est. value stats
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY = '1B2A4A'
WHITE = 'FFFFFF'
ALT = 'EBF5FB'


def _font(sz=10, bold=False, color=NAVY):
    return Font(name='Arial', size=sz, bold=bold, color=color)


def _fill(color=WHITE):
    return PatternFill('solid', fgColor=color)


def _ctr():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _lft(indent=0):
    return Alignment(horizontal='left', vertical='center', indent=indent, wrap_text=True)


def _title_row(ws, row, text, ncols=6, bg=NAVY, fg=WHITE, sz=12, h=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(bg)
    c.font = _font(sz, bold=True, color=fg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2, wrap_text=True)
    ws.row_dimensions[row].height = h


def _section(ws, row, text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = _fill('2C3E50')
    c.font = _font(11, bold=True, color=WHITE)
    c.alignment = _lft(1)
    ws.row_dimensions[row].height = 22


def _txt_block(ws, row, text, ncols=6, bg='F8FBFE', h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = _fill(bg)
    c.font = _font(10)
    c.alignment = _lft(1)
    ws.row_dimensions[row].height = h


def _table_header(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.fill = _fill('344D6E')
        c.font = _font(9, bold=True, color=WHITE)
        c.alignment = _ctr()


def _table_row(ws, row, vals, idx, num_cols=set()):
    bg = ALT if idx % 2 == 0 else WHITE
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row, ci, v)
        c.fill = _fill(bg)
        c.font = _font(10)
        c.alignment = _ctr() if ci in num_cols else _lft(1)


def build_methodology(wb, result):
    """Sheet 0 — full v4 methodology with auto-calibrated weights, factor
    score tables, formula, segmentation cutoffs, and the auto-calibration
    rules that explain why weights change between suburbs."""
    ws = wb.create_sheet('0 - Methodology', 0)
    ws.sheet_view.showGridLines = False

    p = result['profile']
    w = result['weights']
    suburb = result['suburb']
    mh = p.get('median_hold', 0) or 0

    r = 1
    _title_row(ws, r,
        f'📐  Sell Probability Score v4 — Auto-Calibrated Methodology — {suburb}', 6, sz=13, h=30)
    r += 1
    _txt_block(ws, r,
        f"{result['kept_count']:,} properties  |  Median holding: {mh} yrs  |  "
        f"Profile: {'Mature' if p.get('is_mature') else 'Dynamic'} | "
        f"{'Premium' if p.get('is_premium') else 'Standard'} | "
        f"{'High-gain' if p.get('is_high_gain') else 'Moderate-gain'}  |  "
        f"{result.get('today', '')}",
        bg='2E6DA4', h=22)
    ws.cell(r, 1).font = _font(10, bold=True, color=WHITE)
    r += 2

    # Auto-calibrated weights banner
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1,
        f"⚙️  AUTO-CALIBRATED WEIGHTS for {suburb}: "
        f"Hold {w['hold']*100:.0f}% | Type {w['type']*100:.0f}% | "
        f"Gain% {w['gain']*100:.1f}% | CAGR {w['cagr']*100:.1f}% | "
        f"Freq {w['freq']*100:.0f}% | Profit {w['profit']*100:.0f}%")
    c.fill = _fill('E8F4FD')
    c.font = _font(10, bold=True, color='1A5276')
    c.alignment = _lft(1)
    ws.row_dimensions[r].height = 25
    r += 2

    # Factor 1 — Holding
    _section(ws, r, f"FACTOR 1 — Holding Duration   [Weight: {w['hold']*100:.0f}%]")
    r += 1
    _txt_block(ws, r,
        f"Days since current owner purchased ÷ suburb median ({mh} yrs). "
        f"Longer above median = statistically more likely to sell.", h=28)
    r += 1
    _table_header(ws, r, ['Duration', 'Ratio vs Median', 'Score', 'Interpretation', '', ''])
    r += 1
    for i, (d, rt, sc, interp) in enumerate([
        (f'< {mh*0.4:.1f} yrs',           '< 0.4×',     10,  'Well below median — very unlikely'),
        (f'{mh*0.4:.1f}–{mh*0.6:.1f} yrs', '0.4–0.6×',   25,  'Below median — low probability'),
        (f'{mh*0.6:.1f}–{mh*0.8:.1f} yrs', '0.6–0.8×',   40,  'Approaching median'),
        (f'{mh*0.8:.1f}–{mh:.1f} yrs',     '0.8–1.0×',   55,  'Near median — average likelihood'),
        (f'{mh:.1f}–{mh*1.2:.1f} yrs',     '1.0–1.2×',   70,  'Above median — elevated probability'),
        (f'{mh*1.2:.1f}–{mh*1.5:.1f} yrs', '1.2–1.5×',   85,  'Well above median — high probability'),
        (f'> {mh*1.5:.1f} yrs',            '> 1.5×',     100, 'Far above median — very high probability'),
    ], 1):
        _table_row(ws, r, [d, rt, sc, interp, '', ''], i, {1, 2, 3})
        r += 1
    r += 1

    # Factor 2 — Type
    _section(ws, r, f"FACTOR 2 — Property Type   [Weight: {w['type']*100:.0f}%]")
    r += 1
    _table_header(ws, r, ['Type', 'Score', 'Multiplier', 'Rationale', '', ''])
    r += 1
    for i, (t, ts, tm, rat) in enumerate([
        ('House',           100, '×1.3', 'Higher equity, family lifecycle → stronger sell probability'),
        ('Apartment/Unit',   60, '×0.9', 'Often investor-held, lower probability'),
    ], 1):
        _table_row(ws, r, [t, ts, tm, rat, '', ''], i, {1, 2, 3})
        r += 1
    r += 1

    # Factor 3 — Gain%
    if p.get('is_mature') and p.get('is_high_gain'):
        gain_note = (f"PRIMARY signal for {suburb} (mature, high-gain market). The size of the final "
                     f"cheque is what triggers the sale decision; long-holding owners have lower CAGR "
                     f"despite enormous absolute gains.")
    else:
        gain_note = (f"Secondary signal at {w['gain']*100:.1f}% — CAGR is more discriminating in "
                     f"shorter-cycle markets like {suburb}.")
    _section(ws, r, f"FACTOR 3 — Capital Gain %   [Weight: {w['gain']*100:.1f}%]  ← auto-calibrated")
    r += 1
    _txt_block(ws, r, gain_note, h=32)
    r += 1
    _table_header(ws, r, ['Owner Gain %', 'Score', 'Interpretation', '', '', ''])
    r += 1
    for i, (g, gs, gi) in enumerate([
        ('Negative',     15, 'Underwater — very unlikely'),
        ('0–15%',        28, 'Minimal'), ('15–30%', 42, 'Modest'),
        ('30–50%',       55, 'Solid'),   ('50–75%', 68, 'Strong'),
        ('75–100%',      80, 'Very strong'),
        ('100–200%',     90, 'Exceptional'),
        ('200%+',        100, 'Extraordinary — max financial motivation'),
        ('Unknown',      40, 'Neutral — only one sale in dataset'),
    ], 1):
        _table_row(ws, r, [g, gs, gi, '', '', ''], i, {1, 2})
        r += 1
    r += 1

    # Factor 4 — CAGR
    if p.get('is_mature') and p.get('is_high_gain'):
        cagr_note = (f"REFINEMENT signal at {w['cagr']*100:.1f}%. Long-holding premium owners have "
                     f"low CAGR despite huge $ gains; used to penalise short-term anomalies.")
    else:
        cagr_note = (f"PRIMARY performance signal at {w['cagr']*100:.1f}%. Annualised gain is fair "
                     f"across owners with different holding lengths.")
    _section(ws, r, f"FACTOR 4 — CAGR %/yr   [Weight: {w['cagr']*100:.1f}%]  ← auto-calibrated")
    r += 1
    _txt_block(ws, r, cagr_note, h=28)
    r += 1
    _table_header(ws, r, ['CAGR %/yr', 'Score', 'Interpretation', '', '', ''])
    r += 1
    for i, (c_, sc, gi) in enumerate([
        ('Negative', 15, 'Underwater'), ('0–3%', 28, 'Below inflation'),
        ('3–6%',     42, 'Modest'),     ('6–10%', 58, 'Solid'),
        ('10–15%',   72, 'Strong'),     ('15–20%', 88, 'Very strong'),
        ('20%+',     100, 'Exceptional (capped at 50% in data)'),
        ('Unknown',  40, 'Neutral'),
    ], 1):
        _table_row(ws, r, [c_, sc, gi, '', '', ''], i, {1, 2})
        r += 1
    r += 1

    # Factor 5 — Frequency
    _section(ws, r, f"FACTOR 5 — Sale Frequency   [Weight: {w['freq']*100:.0f}%]")
    r += 1
    if p.get('pct_1sale', 0) > 0.35:
        freq_note = (f"Low weight — {p['pct_1sale']*100:.0f}% of {suburb} properties have only "
                     f"one recorded sale. Not enough data to discriminate strongly.")
    else:
        freq_note = (f"More sales = higher liquidity profile = more likely to sell again.")
    _txt_block(ws, r, freq_note, h=24)
    r += 1
    _table_header(ws, r, ['# Historical Sales', 'Score', 'Interpretation', '', '', ''])
    r += 1
    for i, (n, sc, gi) in enumerate([
        ('1', 30, 'No resale history'),
        ('2', 45, 'One resale'),
        ('3', 60, 'Average mobility'),
        ('4', 75, 'Above average'),
        ('5', 88, 'Highly liquid'),
        ('6+', 100, 'Very high turnover'),
    ], 1):
        _table_row(ws, r, [n, sc, gi, '', '', ''], i, {1, 2})
        r += 1
    r += 1

    # Factor 6 — Latent Profit
    _section(ws, r, f"FACTOR 6 — Estimated Latent Profit   [Weight: {w['profit']*100:.0f}%]")
    r += 1
    prof_note = (f"Est. value = land m² × median $/m² from recent sales "
                 f"(House: ${result.get('median_m2_house', 0) or 0:,.0f}/m² | "
                 f"Apt: ${result.get('median_m2_apt', 0) or 0:,.0f}/m²). "
                 f"Profit = Est. value − owner purchase price.")
    _txt_block(ws, r, prof_note, h=32)
    r += 1
    _table_header(ws, r, ['Est. Latent Profit %', 'Score', 'Interpretation', '', '', ''])
    r += 1
    for i, (g, sc, gi) in enumerate([
        ('Negative',  10, 'Underwater'), ('0–20%', 25, 'Minimal'),
        ('20–50%',    40, 'Modest'),     ('50–100%', 58, 'Solid'),
        ('100–200%',  72, 'Strong'),     ('200–400%', 88, 'Very strong'),
        ('400%+',     100, 'Extraordinary'),
        ('Unknown',   40, 'Neutral'),
    ], 1):
        _table_row(ws, r, [g, sc, gi, '', '', ''], i, {1, 2})
        r += 1
    r += 1

    # Formula
    _section(ws, r, 'FORMULA & SEGMENTATION')
    r += 1
    for line in [
        f"Raw = (Hold×{w['hold']*100:.0f}%) + (Type×{w['type']*100:.0f}%) + "
        f"(Gain%×{w['gain']*100:.1f}%) + (CAGR×{w['cagr']*100:.1f}%) + "
        f"(Freq×{w['freq']*100:.0f}%) + (Profit×{w['profit']*100:.0f}%)",
        "Adjusted = Raw × Type Multiplier  (House ×1.3 / Apartment ×0.9)",
        "Final Score = (Adjusted − Min) / (Max − Min) × 100  ← normalised 0–100",
        "",
        f"🔴 HOT    ≥ {result.get('q_hot', 80):.1f}   (top ~18%)",
        f"🟠 WARM   {result.get('q_warm', 60):.1f}–{result.get('q_hot', 80):.1f}   (next ~20%)",
        f"🟡 MEDIUM {result.get('q_medium', 40):.1f}–{result.get('q_warm', 60):.1f}   (monitor)",
        f"⚪ LOW    < {result.get('q_medium', 40):.1f}   (deprioritise)",
    ]:
        _txt_block(ws, r, line, bg='F8FBFE' if line else WHITE, h=18)
        r += 1
    r += 1

    # Auto-calibration rules
    _section(ws, r, 'AUTO-CALIBRATION RULES — Why weights change by suburb')
    r += 1
    for line in [
        "Mature (>20% hold >20yrs) + High-gain (>20% made >200%)  →  Gain% weight ↑↑, CAGR ↓",
        "Dynamic (median hold <10 yrs)  →  CAGR ↑ (shorter cycles = fairer comparison)",
        "Premium (median price >$1M)  →  Latent Profit ↑ (the $ cheque is enormous)",
        "Thin history (>35% single-sale)  →  Frequency ↓ (not enough data to discriminate)",
    ]:
        _txt_block(ws, r, f"  • {line}", bg='EBF5FB', h=22)
        r += 1
    r += 1

    # Warning
    _section(ws, r, '⚠️  Landgate Lag & Agency Note')
    r += 1
    _txt_block(ws, r,
        "Recent sales may not appear for 6–12 months due to Landgate registration delays. "
        "ALWAYS verify on REIWA.com.au or your CRM before prospecting — a HOT lead may have "
        "already sold. Agency/Agent shown = last RECORDED sale, not current.",
        bg='FFF9E6', h=40)

    for col, w_col in zip(['A', 'B', 'C', 'D', 'E', 'F'], [30, 22, 18, 46, 4, 4]):
        ws.column_dimensions[col].width = w_col
    ws.freeze_panes = 'A4'


def build_market_analysis(wb, result):
    """Sheet 2 — Market analysis with suburb profile, weights breakdown, holding
    & gain stats, value/profit stats, and score distribution."""
    ws = wb.create_sheet('2 - Market Analysis')
    ws.sheet_view.showGridLines = False

    p = result['profile']
    w = result['weights']
    suburb = result['suburb']
    properties = result['properties']

    r = 1
    _title_row(ws, r, f'📊  Market Analysis — {suburb}  |  v4 Auto-Calibrated  |  {result.get("today", "")}', 6)
    r += 2

    def block(title, rows):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(r, 1, title)
        c.fill = _fill('2C3E50')
        c.font = _font(10, bold=True, color=WHITE)
        c.alignment = _lft(1)
        ws.row_dimensions[r].height = 20
        r += 1
        for i, (k, v) in enumerate(rows):
            bg = ALT if i % 2 == 0 else WHITE
            kc = ws.cell(r, 1, k)
            vc = ws.cell(r, 2, v)
            kc.fill = vc.fill = _fill(bg)
            kc.font = _font(10, bold=True)
            vc.font = _font(10)
            kc.alignment = vc.alignment = _lft(1)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
        r += 1

    n_hot = sum(1 for x in properties if x['category'] == 'HOT')
    n_warm = sum(1 for x in properties if x['category'] == 'WARM')
    n_med = sum(1 for x in properties if x['category'] == 'MEDIUM')
    n_low = sum(1 for x in properties if x['category'] == 'LOW')
    total = max(len(properties), 1)

    block('Suburb Profile — Auto-Calibration Inputs', [
        ('Classification', f"{'Mature' if p.get('is_mature') else 'Dynamic'} | "
                          f"{'Premium' if p.get('is_premium') else 'Standard'} | "
                          f"{'High-gain' if p.get('is_high_gain') else 'Moderate-gain'}"),
        ('% owners holding > 20 yrs',     f"{p.get('pct_long_hold', 0)*100:.1f}%"),
        ('% with gain > 200%',            f"{p.get('pct_high_gain', 0)*100:.1f}%"),
        ('% single-sale only',            f"{p.get('pct_1sale', 0)*100:.1f}%"),
        ('Median sale price',             f"${(p.get('med_price', 0) or 0):,.0f}"),
        ('Median gain %',                 f"{p.get('median_gain_pct', 0)}%"),
    ])

    block('Auto-Calibrated Weights', [
        ('Holding Duration',         f"{w['hold']*100:.0f}%"),
        ('Property Type',            f"{w['type']*100:.0f}%"),
        ('Capital Gain %',           f"{w['gain']*100:.1f}%"),
        ('CAGR %/yr',                f"{w['cagr']*100:.1f}%"),
        ('Sale Frequency',           f"{w['freq']*100:.0f}%"),
        ('Estimated Latent Profit',  f"{w['profit']*100:.0f}%"),
    ])

    if result.get('rationale'):
        block('Calibration Rationale', [
            ('Why these weights', ', '.join(result['rationale'])),
        ])

    holdings = [x.get('holding_years', 0) or 0 for x in properties]
    holdings_sorted = sorted(holdings)

    def pct(arr, q):
        if not arr:
            return 0
        i = int(len(arr) * q)
        return arr[min(i, len(arr) - 1)]

    block('Holding Period (Last Sale Date → Today)', [
        ('Median (scoring baseline)', f"{p.get('median_hold', 0)} yrs"),
        ('25th percentile',           f"{pct(holdings_sorted, 0.25):.1f} yrs"),
        ('75th percentile',           f"{pct(holdings_sorted, 0.75):.1f} yrs"),
        ('Max',                       f"{max(holdings) if holdings else 0:.1f} yrs"),
    ])

    block('Estimated Value Reference', [
        ('Median $/m² Houses (recent)',     f"${result.get('median_m2_house', 0) or 0:,.0f}"),
        ('Median $/m² Apartments (recent)', f"${result.get('median_m2_apt', 0) or 0:,.0f}"),
    ])

    block('Score Distribution', [
        (f'🔴 HOT (≥ {result.get("q_hot", 80):.1f})',
            f'{n_hot:,} ({n_hot/total*100:.1f}%)'),
        (f'🟠 WARM ({result.get("q_warm", 60):.1f}–{result.get("q_hot", 80):.1f})',
            f'{n_warm:,} ({n_warm/total*100:.1f}%)'),
        (f'🟡 MEDIUM ({result.get("q_medium", 40):.1f}–{result.get("q_warm", 60):.1f})',
            f'{n_med:,} ({n_med/total*100:.1f}%)'),
        (f'⚪ LOW (< {result.get("q_medium", 40):.1f})',
            f'{n_low:,} ({n_low/total*100:.1f}%)'),
        ('High-probability leads (HOT + WARM)', f'{n_hot+n_warm:,} properties'),
        ('Raw records ingested',                f'{result.get("raw_count", 0):,}'),
        ('Excluded by cleaning',                f'{result.get("excluded_count", 0):,}'),
        ('Retained for scoring',                f'{result.get("kept_count", 0):,}'),
    ])

    for col, w_col in zip(['A', 'B'], [42, 28]):
        ws.column_dimensions[col].width = w_col
    ws.freeze_panes = 'A3'
