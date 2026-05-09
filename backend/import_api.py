"""CSV / Excel import for RP Data / CoreLogic property exports.

UPDATE-ONLY mode: rows are matched against existing listings by
(suburb_id, normalized_address). On match, sold_date and sold_price
are overwritten with RP Data values. On no match, the row is skipped.

Performance:
- Pre-fetches every listing in a single query, builds an in-memory
  dict for O(1) match lookup.
- Postgres bulk update uses psycopg2.extras.execute_batch (which
  actually batches at the protocol level, unlike executemany which
  is sneakily one-roundtrip-per-row).
- SQLite executemany is genuinely fast (C-level loop), so dev-mode
  uses that.
"""

import io
import csv
import re
import logging
import time
from datetime import datetime, date
from flask import request, jsonify

from database import get_db, normalize_address, USE_POSTGRES
from admin_api import resolve_request_scope

logger = logging.getLogger(__name__)


COLUMN_ALIASES = {
    'address':      ['property address', 'address', 'street address',
                     'full address', 'street'],
    'suburb':       ['suburb', 'locality', 'area', 'town/suburb'],
    'postcode':     ['postcode', 'post code', 'zip'],
    'sold_date':    ['sale date', 'last sold date', 'date sold',
                     'settlement date', 'sold date', 'transaction date',
                     'contract date'],
    'sold_price':   ['sale price', 'last sold price', 'sold price',
                     'price', 'transaction price'],
    'bedrooms':     ['bedrooms', 'beds', 'bed', 'bedrooms count'],
    'bathrooms':    ['bathrooms', 'baths', 'bath', 'bathrooms count'],
    'parking':      ['car spaces', 'parking', 'cars', 'car', 'garage',
                     'car spaces count'],
    'land_size':    ['land size', 'land area', 'lot size', 'site area',
                     'land (m²)', 'land m²'],
    'internal_size': ['floor area', 'internal area', 'internal size',
                      'building area', 'internal (m²)'],
    'listing_type': ['property type', 'type', 'dwelling type'],
    'agent':        ['agent', 'selling agent', 'sales agent'],
    'agency':       ['agency', 'agency name', 'sales agency',
                     'selling agency'],
    'owner':        ['owner name', 'current owner', 'owner',
                     'registered owner', 'purchaser', 'purchaser name'],
}

RPDATA_POSITIONAL = {
    'address':      0,
    'suburb':       1,
    'postcode':     3,
    'listing_type': 4,
    'bedrooms':     5,
    'bathrooms':    6,
    'parking':      7,
    'land_size':    8,
    'internal_size': 9,
    'sold_price':   11,
    'sold_date':    12,
}

AU_STATE_CODES = {'WA', 'NSW', 'VIC', 'QLD', 'SA', 'TAS', 'NT', 'ACT'}

# Page size for execute_batch — 200 keeps each protocol-level batch
# under ~64KB while still flushing the queue every ~10ms equivalent.
BATCH_PAGE_SIZE = 200


def _parse_price(text):
    if text is None or text == '':
        return None
    s = str(text).strip()
    if not s or s == '-':
        return None
    m_short = re.match(r'^\$?(\d+(?:\.\d+)?)\s*([MmKk])\b', s)
    if m_short:
        try:
            v = float(m_short.group(1))
        except ValueError:
            return None
        suffix = m_short.group(2).lower()
        v *= 1_000_000 if suffix == 'm' else 1_000
        return int(v) if v >= 10_000 else None
    digits = re.sub(r'[^\d.]', '', s)
    if not digits:
        return None
    try:
        v = float(digits)
    except (ValueError, TypeError):
        return None
    return int(v) if v >= 10_000 else None


def _parse_date(text):
    if not text:
        return None
    s = str(text).strip()
    if not s or s == '-':
        return None

    if 'T' in s:
        s = s.split('T', 1)[0]
    parts = s.split(' ', 1)
    if len(parts) > 1 and ':' in parts[1]:
        s = parts[0]

    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            pass

    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1))).isoformat()
        except ValueError:
            pass

    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2})$', s)
    if m:
        try:
            yr = int(m.group(3))
            yr = 2000 + yr if yr < 70 else 1900 + yr
            return date(yr, int(m.group(2)), int(m.group(1))).isoformat()
        except ValueError:
            pass

    for fmt in (
        '%d %b %Y', '%d %B %Y',
        '%b %d %Y', '%B %d %Y',
        '%d-%b-%y', '%d-%b-%Y',
        '%d-%B-%y', '%d-%B-%Y',
    ):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def _is_header_row(cells):
    norm = [(c or '').strip().lower() for c in cells]
    hits = 0
    for aliases in COLUMN_ALIASES.values():
        for alias in aliases:
            if alias in norm:
                hits += 1
                break
    return hits >= 3


def _detect_columns(header):
    norm = [(h or '').strip().lower() for h in header]
    out = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                out[canonical] = norm.index(alias)
                break
    return out


def _looks_like_rpdata_positional(row):
    if not row or len(row) < 5:
        return False
    addr = (row[0] or '').strip()
    suburb = (row[1] or '').strip()
    state = (row[2] or '').strip().upper()
    if not re.match(r'^\d+\w*\s+[A-Z]', addr):
        return False
    if not suburb:
        return False
    if state not in AU_STATE_CODES:
        return False
    return True


def _strip_suburb_from_address(addr, suburb):
    if not addr:
        return addr
    addr = addr.strip().rstrip(',').strip()
    if not suburb:
        return addr
    suburb_lower = suburb.lower()
    addr_lower = addr.lower()
    pos = addr_lower.find(suburb_lower)
    if pos > 0:
        return addr[:pos].strip().rstrip(',').strip()
    return addr


def _read_rows(file_storage):
    name = (file_storage.filename or '').lower()
    if name.endswith('.csv'):
        try:
            text = file_storage.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_storage.seek(0)
            text = file_storage.read().decode('latin-1', errors='replace')
        reader = csv.reader(io.StringIO(text))
        return list(reader), None
    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None, 'openpyxl not installed on the server'
        try:
            wb = load_workbook(file_storage, data_only=True, read_only=True)
        except Exception as e:
            return None, f'Could not open Excel file: {e}'
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append([('' if c is None else str(c)) for c in r])
        return rows, None
    return None, 'Unsupported file extension. Use .csv or .xlsx.'


def _find_header_index(rows, scan_limit=20):
    for i in range(min(scan_limit, len(rows))):
        if _is_header_row(rows[i]):
            return i
    return -1


def import_rpdata():
    """POST /api/listings/import-rpdata.

    Bulk-batched UPDATE-only import. Tuned to handle 2000+ row CSVs
    in under 5 seconds via:
    - 1 SELECT to pre-fetch all listings (in-memory match)
    - psycopg2 execute_batch to bulk-apply UPDATEs at protocol level
    """
    timings = {}
    t0 = time.time()

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded. Use multipart field "file".'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    rows, err = _read_rows(file)
    if err:
        return jsonify({'error': err}), 400
    if not rows:
        return jsonify({'error': 'File is empty'}), 400

    timings['read_file_ms'] = int((time.time() - t0) * 1000)

    header_idx = _find_header_index(rows)
    cols = {}
    data_rows = []
    layout_used = ''

    if header_idx >= 0:
        cols = _detect_columns(rows[header_idx])
        data_rows = rows[header_idx + 1:]
        layout_used = 'named-header'
    elif _looks_like_rpdata_positional(rows[0]):
        cols = dict(RPDATA_POSITIONAL)
        data_rows = rows
        layout_used = 'rpdata-positional'

    if 'address' not in cols:
        return jsonify({
            'error': "Could not parse the file's structure.",
            'first_row_seen': [str(c) for c in rows[0][:10]],
            'header_row_index': header_idx,
        }), 400

    fallback_suburb = (request.form.get('suburb') or '').strip()

    _user, allowed_ids = resolve_request_scope()

    conn = get_db()

    t1 = time.time()
    suburb_rows = conn.execute("SELECT id, name, slug FROM suburbs").fetchall()
    suburb_by_name = {r['name'].lower(): dict(r) for r in suburb_rows}

    listing_rows = conn.execute(
        "SELECT id, suburb_id, normalized_address, sold_date, sold_price, status "
        "FROM listings"
    ).fetchall()
    listings_by_key = {}
    for r in listing_rows:
        norm = (r['normalized_address'] or '').strip().lower()
        if norm:
            listings_by_key[(r['suburb_id'], norm)] = dict(r)
    timings['prefetch_ms'] = int((time.time() - t1) * 1000)

    matched = 0
    no_match = 0
    skipped = 0
    date_updates = 0
    price_updates = 0
    status_updates = 0
    errors = []
    now_iso = datetime.utcnow().isoformat()

    update_payloads = []
    start_offset = (header_idx + 2) if header_idx >= 0 else 1

    t2 = time.time()
    for row_idx, row in enumerate(data_rows, start=start_offset):
        if not row or all((c is None or str(c).strip() in ('', '-')) for c in row):
            continue

        try:
            def cell(key):
                idx = cols.get(key)
                if idx is None or idx >= len(row):
                    return ''
                v = row[idx]
                if v is None:
                    return ''
                s = str(v).strip()
                return '' if s == '-' else s

            addr_raw = cell('address')
            if not addr_raw:
                skipped += 1
                continue

            suburb_name = cell('suburb') or fallback_suburb
            if not suburb_name:
                skipped += 1
                continue

            suburb_row = suburb_by_name.get(suburb_name.lower())
            if not suburb_row:
                no_match += 1
                continue

            suburb_id = suburb_row['id']
            # Multi-tenant scope: silently skip rows whose suburb isn't
            # in the caller's allowed list (admin → allowed_ids None →
            # no filter). Counted as skipped, not errors, so a CSV that
            # straddles agencies imports cleanly for the rows the user
            # owns and ignores the rest without surfacing scope errors.
            if allowed_ids is not None and suburb_id not in allowed_ids:
                skipped += 1
                continue
            address = _strip_suburb_from_address(addr_raw, suburb_name)
            norm_addr = normalize_address(address)
            if not norm_addr:
                skipped += 1
                continue

            existing = listings_by_key.get((suburb_id, norm_addr.lower()))
            if not existing:
                no_match += 1
                continue

            sold_date = _parse_date(cell('sold_date'))
            sold_price = _parse_price(cell('sold_price'))

            new_sold_date = existing['sold_date']
            new_sold_price = existing['sold_price']
            new_status = existing['status']
            changed = False

            if sold_date and sold_date != existing['sold_date']:
                new_sold_date = sold_date
                changed = True
                date_updates += 1
            if sold_price:
                price_str = str(sold_price)
                if price_str != (existing['sold_price'] or ''):
                    new_sold_price = price_str
                    changed = True
                    price_updates += 1
            if existing['status'] != 'sold':
                new_status = 'sold'
                changed = True
                status_updates += 1

            matched += 1
            if changed:
                update_payloads.append((
                    new_sold_date, new_sold_price, new_status,
                    now_iso, existing['id']
                ))
        except Exception as e:
            skipped += 1
            if len(errors) < 20:
                errors.append(f"Row {row_idx}: {type(e).__name__}: {e}")
    timings['match_ms'] = int((time.time() - t2) * 1000)

    # Bulk-apply updates. psycopg2.extras.execute_batch actually batches
    # at the wire-protocol level, unlike cursor.executemany which loops
    # individual UPDATEs and hits the per-row roundtrip wall hard.
    t3 = time.time()
    raw_conn = conn._conn
    cur = raw_conn.cursor()

    if update_payloads:
        if USE_POSTGRES:
            from psycopg2.extras import execute_batch
            update_sql = (
                "UPDATE listings SET sold_date = %s, sold_price = %s, "
                "status = %s, last_seen = %s WHERE id = %s"
            )
            try:
                execute_batch(cur, update_sql, update_payloads,
                              page_size=BATCH_PAGE_SIZE)
            except Exception as e:
                errors.append(f"execute_batch: {type(e).__name__}: {e}")
                try:
                    raw_conn.rollback()
                except Exception:
                    pass
        else:
            update_sql = (
                "UPDATE listings SET sold_date = ?, sold_price = ?, "
                "status = ?, last_seen = ? WHERE id = ?"
            )
            try:
                cur.executemany(update_sql, update_payloads)
            except Exception as e:
                errors.append(f"executemany: {type(e).__name__}: {e}")

    raw_conn.commit()
    conn.close()
    timings['write_ms'] = int((time.time() - t3) * 1000)
    timings['total_ms'] = int((time.time() - t0) * 1000)

    return jsonify({
        'matched': matched,
        'no_match': no_match,
        'skipped': skipped,
        'date_updates': date_updates,
        'price_updates': price_updates,
        'status_updates': status_updates,
        'rows_actually_written': len(update_payloads),
        'total_rows': len(data_rows),
        'layout_used': layout_used,
        'timings_ms': timings,
        'errors': errors,
    })


def register_import_routes(app):
    app.add_url_rule(
        '/api/listings/import-rpdata',
        endpoint='import_rpdata',
        view_func=import_rpdata,
        methods=['POST']
    )
    logger.info("Import routes registered: POST /api/listings/import-rpdata")
